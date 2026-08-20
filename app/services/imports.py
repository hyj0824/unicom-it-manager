from __future__ import annotations

import json
import re
import unicodedata
from pathlib import Path

from sqlalchemy import select
from sqlalchemy.orm import Session

from ..models import BusinessService, ImportBatch, NetworkDevice, StagingRow

# 扁平台账是唯一的人机交换格式：一行表示一个业务与一个设备的关联。
LEDGER_COLUMNS = [
    "号码", "户名", "县分", "网格", "服务状态", "入网时间", "协议到期时间",
    "业务类型", "渠道名称", "发展人", "发展人联系电话", "客户经理",
    "客户经理联系电话", "网络维护责任人", "网络维护责任人联系电话", "设备属性",
    "设备编码", "资产原值或物资购置价格", "设备及物资类型", "设备厂家+型号",
    "设备放置地点", "设备是否已回收", "设备未回收原因",
]
TECHNICAL_COLUMNS = ["业务记录ID", "业务版本", "设备记录ID", "设备版本"]
LEDGER_FIELD_GUIDANCE = [
    ("号码", "必填；业务唯一标识。同一业务有多台设备时，各行号码及业务字段必须保持一致。", "848DIA11742988"),
    ("户名", "必填；填写客户完整名称。", "某某有限公司"),
    ("服务状态", "只能填写：正常开机 / 主动退网(申请拆机)。", "正常开机"),
    ("入网时间", "日期格式：YYYY-MM-DD 或 YYYYMMDD。", "2022-11-03"),
    ("协议到期时间", "日期格式：YYYY-MM-DD 或 YYYYMMDD；未知可留空并进入缺项完善。", "2026-12-31"),
    ("业务类型", "只能填写：数据及网元业务 / 宽带业务。", "宽带业务"),
    ("发展人联系电话", "仅允许数字，可带开头的 +，长度 5-20 位。", "13800000000"),
    ("客户经理联系电话", "仅允许数字，可带开头的 +，长度 5-20 位。", "13800000000"),
    ("网络维护责任人联系电话", "仅允许数字，可带开头的 +，长度 5-20 位。", "13800000000"),
    ("设备属性", "只能填写：资产类 / 成本类。", "资产类"),
    ("设备编码", "推荐一行只填一个设备编码。兼容旧台账多编码：可用顿号、中文/英文逗号、中文/英文分号或换行分隔；不要使用 / 或 |。", "21000001"),
    ("资产原值或物资购置价格", "填写非负数字，不要包含货币单位；千分位逗号可识别。", "1200.00"),
    ("设备是否已回收", "只能填写：已回收 / 未回收；导入时兼容旧值“是 / 否”。", "未回收"),
    ("设备未回收原因", "设备未回收时只能填写：在用 / 遗失 / 纠纷 / 其他；已回收时可留空。", "在用"),
    ("业务记录ID", "系统校验字段，请勿修改。用于准确识别导出时的业务记录。", "系统生成"),
    ("业务版本", "系统校验字段，请勿修改。版本过期时系统会阻止覆盖并提示冲突。", "系统生成"),
    ("设备记录ID", "系统校验字段，请勿修改。新增设备时保持为空。", "系统生成或留空"),
    ("设备版本", "系统校验字段，请勿修改。新增设备时保持为空。", "系统生成或留空"),
]
MISSING_WATCH_COLUMNS = [
    "协议到期时间", "客户经理", "网络维护责任人", "设备属性",
    "资产原值或物资购置价格", "设备放置地点",
]
# 网络维护责任人及联系电话跟随设备记录，进入 network 审批域。
BUSINESS_COLUMNS = LEDGER_COLUMNS[:13]
DEVICE_COLUMNS = LEDGER_COLUMNS[13:]
DEVICE_CODE_SEPARATORS = re.compile(r"[、，,;；\n]+")
PHONE_RE = re.compile(r"^\+?[0-9]{5,20}$")


def _header_key(value: object) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).strip()
    # 说明性括号不属于字段名，例如“设备编码（资产及非资产）”。
    text = re.split(r"[（(]", text, maxsplit=1)[0]
    return re.sub(r"\s+", "", text)


def _headers(path: Path) -> tuple[tuple, list[object]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = list(next(sheet.iter_rows(min_row=2, max_row=2, values_only=True), ()))
    workbook.close()
    return tuple(headers), headers


def header_mapping(path: Path) -> dict[str, str]:
    """Map canonical Chinese columns to unambiguous source headers.

    Matching ignores whitespace and an explanatory parenthesized suffix.  Two
    source columns matching one target are rejected so that the operator, not
    the system, chooses an ambiguous mapping.
    """

    _, headers = _headers(path)
    candidates: dict[str, list[str]] = {}
    for value in headers:
        if value is None or not str(value).strip():
            continue
        candidates.setdefault(_header_key(value), []).append(str(value).strip())
    mapping: dict[str, str] = {}
    missing: list[str] = []
    ambiguous: list[str] = []
    for column in LEDGER_COLUMNS:
        matches = candidates.get(_header_key(column), [])
        if not matches:
            missing.append(column)
        elif len(matches) > 1:
            ambiguous.append(column)
        else:
            mapping[column] = matches[0]
    if missing:
        raise ValueError(f"未识别标准表头：{'、'.join(missing)}")
    if ambiguous:
        raise ValueError(f"表头映射存在多个候选，需人工处理：{'、'.join(ambiguous)}")
    return mapping


def _source_indexes(headers: list[object], mapping: dict[str, str]) -> dict[str, int]:
    indexes: dict[str, int] = {}
    for column, source in mapping.items():
        choices = [i for i, value in enumerate(headers) if str(value or "").strip() == source]
        if len(choices) != 1:
            raise ValueError(f"表头 {source} 无法唯一定位。")
        indexes[column] = choices[0]
    for column in TECHNICAL_COLUMNS:
        choices = [i for i, value in enumerate(headers) if _header_key(value) == _header_key(column)]
        if len(choices) == 1:
            indexes[column] = choices[0]
        elif len(choices) > 1:
            raise ValueError(f"技术列 {column} 存在多个候选。")
    return indexes


def parse_ledger_rows(path: Path) -> list[dict]:
    """Read the active sheet's data rows, preserving the original Chinese values."""

    from openpyxl import load_workbook

    mapping = header_mapping(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = list(next(sheet.iter_rows(min_row=2, max_row=2, values_only=True), ()))
    indexes = _source_indexes(headers, mapping)
    rows: list[dict] = []
    for row_number, values in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
        if all(value is None or str(value).strip() == "" for value in values):
            continue
        raw = {
            column: "" if index >= len(values) or values[index] is None else str(values[index]).strip()
            for column, index in indexes.items()
        }
        for column in LEDGER_COLUMNS + TECHNICAL_COLUMNS:
            raw.setdefault(column, "")
        rows.append({"row_number": row_number, "raw": raw, "source_raw": raw.copy()})
    workbook.close()
    return rows


def _split_device_codes(value: str) -> list[str]:
    return [part.strip() for part in DEVICE_CODE_SEPARATORS.split(value) if part.strip()]


def _int_or_error(value: str, label: str, errors: list[str]) -> int | None:
    if not value:
        return None
    try:
        return int(value)
    except ValueError:
        errors.append(f"{label}必须是整数")
        return None


def _validate_rows(
    rows: list[dict],
    existing_services: dict[str, tuple[int, int]],
    existing_devices: dict[str, tuple[int, int, int]],
) -> list[dict]:
    """Validate a flat ledger. Same-service rows are expected, not duplicates."""

    groups: dict[str, list[dict]] = {}
    device_rows: dict[str, tuple[str, int]] = {}
    for row in rows:
        number = row["raw"].get("号码", "")
        if number:
            groups.setdefault(number, []).append(row)
        for code in _split_device_codes(row["raw"].get("设备编码", "")):
            device_rows.setdefault(code, (number, row["row_number"]))

    result: list[dict] = []
    for row in rows:
        raw = row["raw"]
        errors: list[str] = []
        number = raw.get("号码", "")
        if not number:
            errors.append("业务号码为空")
        if not raw.get("户名"):
            errors.append("客户名称为空")
        for label in ("发展人联系电话", "客户经理联系电话", "网络维护责任人联系电话"):
            if raw.get(label) and not PHONE_RE.match(str(raw[label]).strip()):
                errors.append(f"{label}格式不正确（{raw[label]}）")

        service = existing_services.get(number)
        service_id = _int_or_error(raw.get("业务记录ID", ""), "业务记录ID", errors)
        file_service_version = _int_or_error(raw.get("业务版本", ""), "业务版本", errors)
        if service_id is not None:
            if service is None or service[0] != service_id:
                errors.append("业务记录ID与业务号码不一致或记录不存在")
            else:
                service = (service_id, service[1])
        if service and file_service_version is not None and service[1] != file_service_version:
            errors.append(f"业务记录版本已变化（文件 {file_service_version}，当前 {service[1]}）")

        group = groups.get(number, [])
        if group:
            anchor = group[0]["raw"]
            conflicting = [
                column for column in BUSINESS_COLUMNS
                if raw.get(column, "") != anchor.get(column, "")
            ]
            if conflicting:
                errors.append(f"同一业务号码的业务字段不一致：{'、'.join(conflicting)}")

        device_codes = _split_device_codes(raw.get("设备编码", ""))
        if re.search(r"[/|]", raw.get("设备编码", "")):
            errors.append("设备编码包含未支持的分隔符（/ 或 |），请在缺项工作台拆分确认")
        if len(device_codes) != len(set(device_codes)):
            errors.append("同一行重复填写了设备编码")
        has_device_values = any(raw.get(column, "") for column in LEDGER_COLUMNS[15:] if column != "设备编码")
        if len(group) > 1 and not device_codes and any(_split_device_codes(item["raw"].get("设备编码", "")) for item in group):
            errors.append(f"业务号码 {number} 在本批次内重复且该行未提供设备编码")
        devices = []
        for code in device_codes:
            first_number, first_row = device_rows[code]
            if first_row != row["row_number"]:
                errors.append(
                    f"设备编码 {code} 同时出现在多个业务行（首次出现在第 {first_row} 行）"
                )
            existing = existing_devices.get(code)
            device_id = _int_or_error(raw.get("设备记录ID", ""), "设备记录ID", errors)
            file_device_version = _int_or_error(raw.get("设备版本", ""), "设备版本", errors)
            if device_id is not None:
                if existing is None or existing[1] != device_id:
                    errors.append(f"设备记录ID与设备编码 {code} 不一致或记录不存在")
                else:
                    existing = (existing[0], device_id, existing[2])
            if existing and existing[0] != (service[0] if service else None):
                errors.append(f"设备编码 {code} 已归属其他业务，不能跨业务覆盖")
            if existing and file_device_version is not None and existing[2] != file_device_version:
                errors.append(f"设备编码 {code} 的记录版本已变化（文件 {file_device_version}，当前 {existing[2]}）")
            devices.append({
                "device_code": code,
                "existing_device_id": existing[1] if existing else None,
                "base_version": existing[2] if existing else None,
                "asset_class": raw.get("设备属性", ""),
                "asset_value": raw.get("资产原值或物资购置价格", ""),
                "device_type": raw.get("设备及物资类型", ""),
                "vendor_model": raw.get("设备厂家+型号", ""),
                "location": raw.get("设备放置地点", ""),
                "recovery_status": raw.get("设备是否已回收", ""),
                "recovery_reason": raw.get("设备未回收原因", ""),
                "maintenance_name": raw.get("网络维护责任人", ""),
                "maintenance_phone": raw.get("网络维护责任人联系电话", ""),
            })
        mapped = {
            "operation": "update" if service else "create",
            "existing_service_id": service[0] if service else None,
            "base_version": service[1] if service else None,
            "service_number": number,
            "customer_name": raw.get("户名", ""),
            "county": raw.get("县分", ""), "grid": raw.get("网格", ""),
            "service_status": raw.get("服务状态", ""), "business_type": raw.get("业务类型", ""),
            "channel_name": raw.get("渠道名称", ""), "accessed_at": raw.get("入网时间", ""),
            "agreement_expires_at": raw.get("协议到期时间", ""),
            "developer_name": raw.get("发展人", ""),
            "developer_phone": raw.get("发展人联系电话", ""),
            "account_manager_name": raw.get("客户经理", ""),
            "account_manager_phone": raw.get("客户经理联系电话", ""),
            "devices": devices,
            "source_raw": row.get("source_raw", raw),
        }
        if raw != row.get("source_raw", raw):
            mapped["manual_raw"] = raw
        missing = [column for column in MISSING_WATCH_COLUMNS if not raw.get(column)]
        if has_device_values and not device_codes:
            missing.append("设备编码")
        status = "error" if errors else "missing" if missing else "valid"
        result.append({
            "row_number": row["row_number"],
            "raw_json": json.dumps(row.get("source_raw", raw), ensure_ascii=False),
            "mapped_json": json.dumps(mapped, ensure_ascii=False),
            "status": status,
            "error_messages": "\n".join(errors),
        })
    return result


def _existing_maps(db: Session):
    services = {
        service.service_number: (service.id, service.version)
        for service in db.scalars(select(BusinessService)).all()
    }
    devices = {
        device.device_code: (device.business_service_id, device.id, device.version)
        for device in db.scalars(select(NetworkDevice)).all()
    }
    return services, devices


def _set_batch_counts(batch: ImportBatch, validated: list[dict]) -> None:
    batch.total_rows = len(validated)
    batch.error_count = sum(row["status"] == "error" for row in validated)
    batch.missing_count = sum(row["status"] == "missing" for row in validated)
    batch.duplicate_count = sum("多个业务行" in row["error_messages"] or "其他业务" in row["error_messages"] for row in validated)


def import_ledger_workbook(db: Session, batch: ImportBatch, path: Path) -> ImportBatch:
    rows = parse_ledger_rows(path)
    services, devices = _existing_maps(db)
    validated = _validate_rows(rows, services, devices)
    _set_batch_counts(batch, validated)
    batch.header_mapping_json = json.dumps(header_mapping(path), ensure_ascii=False)
    for row in validated:
        db.add(StagingRow(batch=batch, **row))
    batch.status = "ready"
    return batch


def revalidate_staging_row(db: Session, batch: ImportBatch, row_id: int, corrected: dict[str, str]) -> StagingRow:
    """Keep original upload values, apply an operator correction, and revalidate the batch."""

    target = db.get(StagingRow, row_id)
    if target is None or target.batch_id != batch.id:
        raise ValueError("暂存行不存在。")
    rows: list[dict] = []
    for staging in db.scalars(select(StagingRow).where(StagingRow.batch_id == batch.id)).all():
        mapped = json.loads(staging.mapped_json or "{}")
        source = json.loads(staging.raw_json or "{}")
        raw = mapped.get("manual_raw", source)
        if staging.id == row_id:
            raw = {column: str(corrected.get(column, "")).strip() for column in LEDGER_COLUMNS + TECHNICAL_COLUMNS}
        rows.append({"row_number": staging.row_number, "raw": raw, "source_raw": source, "id": staging.id})
    services, devices = _existing_maps(db)
    validated = _validate_rows(rows, services, devices)
    by_number = {row["row_number"]: row for row in validated}
    for staging in db.scalars(select(StagingRow).where(StagingRow.batch_id == batch.id)).all():
        value = by_number[staging.row_number]
        staging.raw_json = value["raw_json"]
        staging.mapped_json = value["mapped_json"]
        staging.status = value["status"]
        staging.error_messages = value["error_messages"]
    _set_batch_counts(batch, validated)
    batch.status = "ready"
    return target
