from __future__ import annotations

import hashlib
import io
import json
import uuid
from contextlib import asynccontextmanager
from datetime import datetime, time, timedelta, timezone
from pathlib import Path
from urllib.parse import quote

from fastapi import Depends, FastAPI, HTTPException, Request, UploadFile, status
from fastapi.responses import FileResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy import case, func, or_, select, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session
from starlette.middleware.sessions import SessionMiddleware

from . import auth
from .audio import resolve_audio_file
from .config import (
    BASE_DIR,
    ensure_storage_paths,
    get_settings,
    validate_runtime_settings,
)
from .database import SessionLocal, check_schema_current, get_db, get_schema_head
from .logging import configure_logging
from .models import (
    AuditLog,
    CALL_STATUSES,
    CallEvent,
    CallRecord,
    CallTask,
    ChangeSet,
    Contact,
    Customer,
    BusinessService,
    DictionaryItem,
    ImportBatch,
    NetworkDevice,
    Role,
    RolePermission,
    ScanSchedule,
    Script,
    SmsNotification,
    StagingRow,
    User,
    UserRole,
    utcnow,
)
from .scheduler import scheduler_service
from .services import imports as import_service
from .services import ledger as ledger_service
from .services import reviews as review_service
from .services import change_requests as change_request_service
from .services import plans as plan_service
from .services.call_worker import call_worker_service, modem_availability
from .services.backups import backup_service
from .services.dictionaries import active_items, resolve_or_create_item
from .services.scripts import (
    generate_script_audio,
    script_audio_url,
    referencing_counts as script_referencing_counts,
)
from .services.settings import (
    SCHEDULER_ENABLED_KEY,
    CALL_WORKER_ENABLED_KEY,
    ensure_default_settings,
    is_scheduler_enabled,
    is_worker_enabled,
    set_setting,
)
from .services.users import (
    hash_password,
    role_names,
    set_user_roles,
    validate_user_profile,
    verify_password,
)


settings = get_settings()
configure_logging(
    redact_secrets=(settings.admin_password, settings.session_secret, settings.tts_api_key)
)
templates = Jinja2Templates(directory=str(Path(__file__).parent / "templates"))

APP_TITLE = "中国联通 IT 运维客户信息管理系统"

STATUS_LABELS = {
    "uploaded": "已上传",
    "validating": "校验中",
    "ready": "待提交",
    "reviewing": "审核中",
    "applied": "已应用",
    "rejected": "已驳回",
    "draft": "草稿",
    "submitted": "待审核",
    "returned": "已退回",
    "approved": "已通过",
    "cancelled": "已取消",
    "valid": "正常",
    "missing": "缺项",
    "error": "错误",
    "duplicate": "冲突",
    "queued": "待拨打",
    "dialing": "拨号中",
    "connected": "已接通",
    "no_answer": "无人接听",
    "rejected": "疑似拒接",
    "cancelled_or_failed": "已取消或失败",
    "busy": "忙线",
    "short_call": "有效时长不足",
    "failed": "失败",
    "completed": "已完成",
    "missed": "已错过",
}

DOMAIN_LABELS = {
    "business": "业务",
    "network": "网络",
    "callback": "回访",
    "template": "话术",
    "system": "系统",
}

# 扫描通知配置的扫描类型（app/models.py ScanSchedule.scan_type）。
SCAN_TYPE_LABELS = {
    "due_renewal": "到期维系",
    "device_recycle": "设备回收",
    "review_stuck": "审核卡单",
}

# 通知任务来源（CallTask.source）；历史 callback_plans 生成的任务仍为 scheduled。
SOURCE_LABELS = {
    "manual": "立即拨打",
    "due_renewal": "到期维系扫描",
    "device_recycle": "设备回收扫描",
    "review_stuck": "审核卡单提醒扫描",
    "scheduled": "计划触发",
}


# 短信通知状态（app/models.py SmsNotification.status）。
SMS_STATUS_LABELS = {
    "pending": "待发送",
    "sent": "已发送",
    "failed": "失败",
}


def mask_phone(phone: str) -> str:
    """手机号脱敏：138****0000；过短或异常号码原样展示。"""

    phone = (phone or "").strip()
    if len(phone) < 8:
        return phone
    return phone[:3] + "*" * (len(phone) - 7) + phone[-4:]


def sms_status_label(value: str) -> str:
    return SMS_STATUS_LABELS.get(value, value or "-")


def _as_utc(value: datetime | None) -> datetime | None:
    if value is None:
        return None
    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc)


def format_dt(value: datetime | None, timezone_name: str | None = None) -> str:
    value = _as_utc(value)
    if value is None:
        return "-"
    zone = plan_service.get_zone(timezone_name or settings.default_timezone)
    return value.astimezone(zone).strftime("%Y-%m-%d %H:%M")


def format_dt_seconds(value: datetime | None, timezone_name: str | None = None) -> str:
    """带秒的时间格式，用于通话详情事件时间线。

    事件在创建时盖章（见 `CallEvent.__init__`），实测关键事件间隔为秒级
    （拨出→响铃 1.3s、响铃期 13.9s 等），秒级显示即可分辨。
    """
    value = _as_utc(value)
    if value is None:
        return "-"
    zone = plan_service.get_zone(timezone_name or settings.default_timezone)
    return value.astimezone(zone).strftime("%Y-%m-%d %H:%M:%S")


def format_date(value: datetime | None, timezone_name: str | None = None) -> str:
    value = _as_utc(value)
    if value is None:
        return ""
    zone = plan_service.get_zone(timezone_name or settings.default_timezone)
    return value.astimezone(zone).strftime("%Y-%m-%d")


def html_date_value(value: object) -> str:
    """Normalize imported date text for an HTML date input."""

    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    if len(text) == 8 and text.isdigit():
        return f"{text[:4]}-{text[4:6]}-{text[6:]}"
    if len(text) >= 10 and text[4] == "-" and text[7] == "-":
        return text[:10]
    return text


def from_json_map(value: str, key: str) -> str:
    try:
        data = json.loads(value or "{}")
    except ValueError:
        return "-"
    result = data.get(key, "")
    return str(result) if result else "-"


def status_label(value: str) -> str:
    return STATUS_LABELS.get(value, value or "-")


def domain_label(value: str) -> str:
    return DOMAIN_LABELS.get(value, value or "-")


def source_label(value: str) -> str:
    return SOURCE_LABELS.get(value, value or "-")


def scan_type_label(value: str) -> str:
    return SCAN_TYPE_LABELS.get(value, value or "-")


templates.env.filters["dt"] = format_dt
templates.env.filters["dt_s"] = format_dt_seconds
templates.env.filters["date"] = format_date
templates.env.filters["from_json_map"] = from_json_map
templates.env.filters["status_label"] = status_label
templates.env.filters["domain_label"] = domain_label
templates.env.filters["source_label"] = source_label
templates.env.filters["scan_type_label"] = scan_type_label
templates.env.filters["mask_phone"] = mask_phone
templates.env.filters["sms_status_label"] = sms_status_label
templates.env.filters["script_audio_url"] = script_audio_url


@asynccontextmanager
async def lifespan(app: FastAPI):
    validate_runtime_settings(settings)
    ensure_storage_paths(settings)
    check_schema_current()
    (BASE_DIR / "data" / "imports").mkdir(parents=True, exist_ok=True)
    with SessionLocal() as db:
        ensure_default_settings(db)
        db.commit()
    scheduler_service.start()
    call_worker_service.start()
    backup_service.start()
    yield
    backup_service.shutdown()
    call_worker_service.shutdown()
    scheduler_service.shutdown()


app = FastAPI(title=APP_TITLE, lifespan=lifespan)
app.add_middleware(SessionMiddleware, secret_key=settings.cookie_secret, same_site="lax")
app.mount("/static", StaticFiles(directory=str(Path(__file__).parent / "static")), name="static")


def redirect_to(path: str) -> RedirectResponse:
    return RedirectResponse(path, status_code=status.HTTP_303_SEE_OTHER)


def context(request: Request, db: Session | None = None, **extra):
    data = {
        "request": request,
        "is_authenticated": auth.is_authenticated(request),
        "current_user": auth.current_user(request),
        "scheduler_enabled": True,
        "settings": settings,
        "app_title": APP_TITLE,
    }
    if db is not None:
        ensure_default_settings(db)
        data["scheduler_enabled"] = is_scheduler_enabled(db)
    data.update(extra)
    return data


def render(
    request: Request,
    template_name: str,
    db: Session | None = None,
    status_code: int = 200,
    **extra,
):
    return templates.TemplateResponse(
        request=request,
        name=template_name,
        context=context(request, db, **extra),
        status_code=status_code,
    )


def get_or_404(db: Session, model, item_id: int):
    item = db.get(model, item_id)
    if item is None:
        raise HTTPException(status_code=404)
    return item


def _form_int(form, key: str) -> int | None:
    value = str(form.get(key, "")).strip()
    if not value:
        return None
    try:
        return int(value)
    except ValueError:
        return None


def _form_float(form, key: str) -> float | None:
    value = str(form.get(key, "")).strip()
    if not value:
        return None
    try:
        return float(value)
    except ValueError:
        raise ValueError(f"{key} 必须是数字。") from None


def _dictionary_value_id(db: Session, category: str, value: str) -> int | None:
    item = resolve_or_create_item(db, category, value)
    return item.id if item is not None else None


def _user_id(request: Request) -> int | None:
    principal = auth.current_user(request)
    if principal and principal.get("type") == "user":
        return principal.get("id")
    return None


def _is_self_submission(request: Request, change_set: ChangeSet) -> bool:
    principal = auth.current_user(request)
    if principal is None:
        return False
    if principal.get("type") == "admin":
        return change_set.created_by_user_id is None
    return change_set.created_by_user_id == _user_id(request)


def _client_ip(request: Request) -> str:
    return request.client.host if request.client else ""


@app.get("/healthz")
def healthz():
    return {"ok": True}


# ---------------------------------------------------------------- 登录与审计


@app.get("/login")
def login_page(request: Request):
    if auth.is_authenticated(request):
        return redirect_to("/")
    return render(request, "login.html")


@app.post("/login")
async def login_submit(request: Request, db: Session = Depends(get_db)):
    form = await request.form()
    username = str(form.get("username", ""))
    password = str(form.get("password", ""))
    principal = auth.verify_credentials(db, username, password)
    if principal is None:
        ledger_service.log_action(
            db, "login_failed", detail=json.dumps({"username": username}, ensure_ascii=False),
            ip_address=_client_ip(request),
        )
        db.commit()
        return render(
            request,
            "login.html",
            status_code=status.HTTP_401_UNAUTHORIZED,
            error="用户名或密码不正确。",
        )
    auth.login(request, principal)
    ledger_service.log_action(
        db,
        "login",
        _user_id(request),
        detail=json.dumps({"username": username}, ensure_ascii=False),
        ip_address=_client_ip(request),
    )
    db.commit()
    if principal.get("type") == "user":
        user = db.get(User, principal.get("id"))
        if user is not None and user.force_password_change:
            # 自动建号账号初始密码为随机值，首次登录必须先改密。
            return redirect_to("/password-change")
    return redirect_to("/")


@app.post("/logout")
def logout_submit(request: Request):
    auth.logout(request)
    return redirect_to("/login")


# ---------------------------------------------------------------- 首次登录改密


@app.get("/password-change")
def password_change_page(request: Request):
    auth.require_login(request)
    return render(request, "password_change.html")


@app.post("/password-change")
async def password_change_submit(request: Request, db: Session = Depends(get_db)):
    auth.require_login(request)
    principal = auth.current_user(request)
    if principal is None or principal.get("type") != "user":
        return redirect_to("/")
    user = db.get(User, principal.get("id"))
    if user is None:
        return redirect_to("/")
    form = await request.form()
    old_password = str(form.get("old_password", ""))
    new_password = str(form.get("new_password", ""))
    confirm_password = str(form.get("confirm_password", ""))
    error = None
    if not verify_password(old_password, user.password_hash):
        error = "当前密码不正确。"
    elif len(new_password) < 8:
        error = "新密码至少 8 位。"
    elif new_password != confirm_password:
        error = "两次输入的新密码不一致。"
    if error:
        return render(request, "password_change.html", error=error, status_code=400)
    was_forced = user.force_password_change
    user.password_hash = hash_password(new_password)
    user.force_password_change = False
    ledger_service.log_action(
        db,
        "change_password",
        _user_id(request),
        "user",
        user.id,
        json.dumps({"forced_first_login": was_forced}, ensure_ascii=False),
        _client_ip(request),
    )
    db.commit()
    return redirect_to("/")


# ---------------------------------------------------------------- 工作台


@app.get("/")
def dashboard(request: Request, db: Session = Depends(get_db)):
    auth.require_login(request)
    now = utcnow()
    expiring_soon = now + timedelta(days=30)

    active_task = db.scalars(
        select(CallTask)
        .where(CallTask.status.in_(["queued", "dialing", "connected"]))
        .order_by(CallTask.due_at.asc(), CallTask.created_at.asc())
        .limit(1)
    ).first()
    pending_reviews = db.scalar(
        select(func.count(ChangeSet.id)).where(ChangeSet.status == "submitted")
    ) or 0
    missing_total = sum(row["count"] for row in ledger_service.business_missing_fields(db))  # type: ignore[arg-type]
    missing_total += sum(row["count"] for row in ledger_service.device_missing_fields(db))  # type: ignore[arg-type]
    expiring_services = db.scalar(
        select(func.count(BusinessService.id)).where(
            BusinessService.is_active.is_(True),
            BusinessService.agreement_expires_at.is_not(None),
            BusinessService.agreement_expires_at <= expiring_soon,
        )
    ) or 0

    call_stats = {
        name: db.scalar(select(func.count(CallRecord.id)).where(CallRecord.status == name)) or 0
        for name in ["completed", "failed", "short_call", "no_answer"]
    }
    counts = {
        "services": db.scalar(
            select(func.count(BusinessService.id)).where(BusinessService.is_active.is_(True))
        ) or 0,
        "devices": db.scalar(
            select(func.count(NetworkDevice.id)).where(NetworkDevice.is_active.is_(True))
        ) or 0,
        "customers": db.scalar(select(func.count(Customer.id))) or 0,
        "scan_schedules": db.scalar(select(func.count(ScanSchedule.id))) or 0,
        "records": db.scalar(select(func.count(CallRecord.id))) or 0,
    }
    scan_enabled = db.scalar(
        select(func.count(ScanSchedule.id)).where(ScanSchedule.enabled.is_(True))
    ) or 0
    recent_records = db.scalars(
        select(CallRecord).order_by(CallRecord.created_at.desc()).limit(8)
    ).all()
    recent_batches = db.scalars(
        select(ImportBatch)
        .where(ImportBatch.error_count > 0)
        .order_by(ImportBatch.created_at.desc())
        .limit(5)
    ).all()
    # 仪表盘 Worker 状态：硬开关（.env）、运行时开关、进程状态、当前通话，
    # 以及串口可用性的只读判断（不打开串口、不发送 AT 指令）。
    worker_status = call_worker_service.status()
    return render(
        request,
        "dashboard.html",
        db,
        active_task=active_task,
        pending_reviews=pending_reviews,
        missing_total=missing_total,
        expiring_services=expiring_services,
        scan_enabled=scan_enabled,
        call_stats=call_stats,
        counts=counts,
        recent_records=recent_records,
        recent_batches=recent_batches,
        worker_status=worker_status,
        worker_gate=settings.call_worker_enabled,
        worker_enabled=is_worker_enabled(db),
        modem_status=modem_availability(settings, worker_status),
    )


@app.post("/settings/scheduler")
async def scheduler_toggle(request: Request, db: Session = Depends(get_db)):
    auth.require_permission(db, request, "manage_config", "system")
    form = await request.form()
    enabled = str(form.get("enabled", "0")) == "1"
    set_setting(db, SCHEDULER_ENABLED_KEY, "1" if enabled else "0")
    db.commit()
    return redirect_to(str(form.get("next", "/")))


@app.post("/settings/worker")
async def worker_toggle(request: Request, db: Session = Depends(get_db)):
    auth.require_permission(db, request, "manage_config", "system")
    form = await request.form()
    enabled = str(form.get("enabled", "0")) == "1"
    call_worker_service.set_enabled(db, enabled)
    db.commit()
    return redirect_to(str(form.get("next", "/admin/system")))


# ---------------------------------------------------------------- 通讯录


def contacts_template(
    request: Request,
    db: Session,
    error: str = "",
    form_data: dict[str, str] | None = None,
    edit_contact: Contact | None = None,
    status_code: int = 200,
):
    contacts = db.scalars(
        select(Contact).where(Contact.is_active.is_(True)).order_by(Contact.name.asc(), Contact.id.asc())
    ).all()
    return render(
        request,
        "contacts.html",
        db,
        status_code=status_code,
        contacts=contacts,
        duties=active_items(db, "contact_duty"),
        edit_contact=edit_contact,
        form_data=form_data or {},
        error=error,
    )


@app.get("/contacts")
def contacts_page(request: Request, edit_id: int | None = None, db: Session = Depends(get_db)):
    auth.require_login(request)
    return contacts_template(request, db, edit_contact=db.get(Contact, edit_id) if edit_id else None)


@app.post("/contacts")
async def directory_contact_create(request: Request, db: Session = Depends(get_db)):
    auth.require_login(request)
    form = await request.form()
    name = str(form.get("name", "")).strip()
    phone = str(form.get("phone", "")).strip()
    duty = str(form.get("duty", "")).strip()
    if not name or not phone:
        return contacts_template(request, db, "姓名和联系电话不能为空。", dict(form), status_code=400)
    if not plan_service.validate_phone(phone):
        return contacts_template(request, db, "电话格式不正确（应为 +?[0-9]{5,20}）。", dict(form), status_code=400)
    db.add(Contact(name=name, phone=phone, duty=duty))
    db.commit()
    return redirect_to("/contacts")


@app.post("/contacts/{contact_id}/edit")
async def contact_update(contact_id: int, request: Request, db: Session = Depends(get_db)):
    auth.require_login(request)
    contact = get_or_404(db, Contact, contact_id)
    form = await request.form()
    name = str(form.get("name", "")).strip()
    phone = str(form.get("phone", "")).strip()
    duty = str(form.get("duty", "")).strip()
    form_data = {"name": name, "phone": phone, "duty": duty}
    if not name or not phone:
        return contacts_template(request, db, "姓名和联系电话不能为空。", form_data, edit_contact=contact, status_code=400)
    if phone and not plan_service.validate_phone(phone):
        return contacts_template(request, db, "电话格式不正确（应为 +?[0-9]{5,20}）。", form_data, edit_contact=contact, status_code=400)
    contact.name = name or None
    contact.phone = phone or None
    contact.duty = duty
    contact.version += 1
    db.commit()
    return redirect_to("/contacts")


@app.post("/contacts/{contact_id}/delete")
def contact_delete(contact_id: int, request: Request, db: Session = Depends(get_db)):
    auth.require_login(request)
    contact = get_or_404(db, Contact, contact_id)
    contact.is_active = False
    db.commit()
    return redirect_to("/contacts")


# ---------------------------------------------------------------- 业务台账


def ledger_template(
    request: Request,
    db: Session,
    error: str = "",
    form_data: dict[str, str] | None = None,
    edit_service: BusinessService | None = None,
    status_code: int = 200,
    q: str = "",
    county_id: int | None = None,
    grid_id: int | None = None,
    status_id: int | None = None,
    business_type_id: int | None = None,
    page: int = 1,
):
    query = select(BusinessService).where(BusinessService.is_active.is_(True))
    if q:
        like = f"%{q.strip()}%"
        query = query.join(Customer, BusinessService.customer_id == Customer.id).where(
            or_(BusinessService.service_number.ilike(like), Customer.name.ilike(like))
        )
    if county_id:
        query = query.where(BusinessService.county_item_id == county_id)
    if grid_id:
        query = query.where(BusinessService.grid_item_id == grid_id)
    if status_id:
        query = query.where(BusinessService.service_status_item_id == status_id)
    if business_type_id:
        query = query.where(BusinessService.business_type_item_id == business_type_id)
    total = db.scalar(select(func.count()).select_from(query.order_by(None).subquery())) or 0
    page_size = 50
    last_page = max(1, (total + page_size - 1) // page_size)
    page = max(1, min(page, last_page))
    services = db.scalars(
        query.order_by(BusinessService.created_at.desc(), BusinessService.id.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
    ).all()
    customers = db.scalars(select(Customer).order_by(Customer.name.asc())).all()
    return render(
        request,
        "ledger.html",
        db,
        status_code=status_code,
        services=services,
        customers=customers,
        counties=active_items(db, "county"),
        grids=active_items(db, "grid"),
        service_statuses=active_items(db, "service_status"),
        business_types=active_items(db, "business_type"),
        edit_service=edit_service,
        form_data=form_data or {},
        q=q,
        county_id=county_id or 0,
        grid_id=grid_id or 0,
        status_id=status_id or 0,
        business_type_id=business_type_id or 0,
        page=page,
        total=total,
        last_page=last_page,
        error=error,
    )


@app.get("/ledger")
def ledger_page(
    request: Request,
    edit_id: int | None = None,
    q: str = "",
    county_id: int | None = None,
    grid_id: int | None = None,
    status_id: int | None = None,
    business_type_id: int | None = None,
    page: int = 1,
    error: str = "",
    db: Session = Depends(get_db),
):
    auth.require_login(request)
    edit_service = db.get(BusinessService, edit_id) if edit_id else None
    return ledger_template(
        request, db, edit_service=edit_service, q=q,
        county_id=county_id, grid_id=grid_id, status_id=status_id,
        business_type_id=business_type_id, page=page, error=error,
    )


def _parse_service_form(form) -> dict:
    return {
        "service_number": str(form.get("service_number", "")).strip(),
        "customer_id": _form_int(form, "customer_id"),
        "county": str(form.get("county", "")).strip(),
        "grid": str(form.get("grid", "")).strip(),
        "service_status": str(form.get("service_status", "")).strip(),
        "business_type": str(form.get("business_type", "")).strip(),
        "channel_name": str(form.get("channel_name", "")).strip(),
        "accessed_at": ledger_service.parse_local_date(
            str(form.get("accessed_at", "")), settings.default_timezone
        ),
        "agreement_expires_at": ledger_service.parse_local_date(
            str(form.get("agreement_expires_at", "")), settings.default_timezone
        ),
    }


def _apply_service_form(service: BusinessService, data: dict, db: Session) -> str | None:
    error = ledger_service.validate_service_number(
        db, data["service_number"], exclude_id=service.id
    )
    if error:
        return error
    if data["customer_id"] is None:
        return "必须选择客户。"
    service.service_number = data["service_number"]
    service.customer_id = data["customer_id"]
    service.county_item_id = _dictionary_value_id(db, "county", data["county"])
    service.grid_item_id = _dictionary_value_id(db, "grid", data["grid"])
    service.service_status_item_id = _dictionary_value_id(db, "service_status", data["service_status"])
    service.business_type_item_id = _dictionary_value_id(db, "business_type", data["business_type"])
    service.channel_name = data["channel_name"]
    service.accessed_at = data["accessed_at"]
    service.agreement_expires_at = data["agreement_expires_at"]
    service.version = (service.version or 0) + 1
    return None


@app.post("/ledger")
async def service_create(request: Request, db: Session = Depends(get_db)):
    auth.require_login(request)
    form = await request.form()
    try:
        data = _parse_service_form(form)
        service = BusinessService(
            service_number="", customer_id=data["customer_id"] or 0
        )
        error = _apply_service_form(service, data, db)
    except ValueError as exc:
        return redirect_to(f"/ledger?error={quote(str(exc))}")
    if error:
        return redirect_to(f"/ledger?error={quote(error)}")
    db.add(service)
    db.commit()
    ledger_service.log_action(
        db, "change", _user_id(request), "business_service", service.id,
        json.dumps({"operation": "create", "service_number": service.service_number}, ensure_ascii=False),
        _client_ip(request),
    )
    db.commit()
    return redirect_to("/ledger")


@app.post("/ledger/{service_id}/edit")
async def service_update(service_id: int, request: Request, db: Session = Depends(get_db)):
    auth.require_login(request)
    service = get_or_404(db, BusinessService, service_id)
    form = await request.form()
    try:
        data = _parse_service_form(form)
        error = _apply_service_form(service, data, db)
    except ValueError as exc:
        return redirect_to(f"/ledger?error={quote(str(exc))}")
    if error:
        return redirect_to(f"/ledger?error={quote(error)}")
    db.commit()
    ledger_service.log_action(
        db, "change", _user_id(request), "business_service", service.id,
        json.dumps({"operation": "update", "service_number": service.service_number}, ensure_ascii=False),
        _client_ip(request),
    )
    db.commit()
    return redirect_to("/ledger")


@app.post("/ledger/{service_id}/delete")
def service_delete(service_id: int, request: Request, db: Session = Depends(get_db)):
    auth.require_login(request)
    service = get_or_404(db, BusinessService, service_id)
    if service.devices:
        return ledger_template(
            request,
            db,
            f"该业务下有 {len(service.devices)} 台设备，请先处理设备关联，不能删除。",
            status_code=400,
        )
    service.is_active = False
    db.commit()
    return redirect_to("/ledger")


# ---------------------------------------------------------------- 网络设备


def devices_template(
    request: Request,
    db: Session,
    error: str = "",
    form_data: dict[str, str] | None = None,
    edit_device: NetworkDevice | None = None,
    status_code: int = 200,
    q: str = "",
    recovery_status_id: int | None = None,
    page: int = 1,
):
    query = select(NetworkDevice).where(NetworkDevice.is_active.is_(True))
    if q:
        like = f"%{q.strip()}%"
        query = query.where(
            or_(NetworkDevice.device_code.ilike(like), NetworkDevice.vendor_model.ilike(like))
        )
    if recovery_status_id:
        query = query.where(NetworkDevice.recovery_status_item_id == recovery_status_id)
    total = db.scalar(select(func.count()).select_from(query.order_by(None).subquery())) or 0
    page_size = 50
    last_page = max(1, (total + page_size - 1) // page_size)
    page = max(1, min(page, last_page))
    devices = db.scalars(
        query.order_by(NetworkDevice.created_at.desc(), NetworkDevice.id.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
    ).all()
    services = db.scalars(
        select(BusinessService)
        .where(BusinessService.is_active.is_(True))
        .order_by(BusinessService.service_number.asc())
    ).all()
    return render(
        request,
        "devices.html",
        db,
        status_code=status_code,
        devices=devices,
        services=services,
        asset_classes=active_items(db, "asset_class"),
        device_types=active_items(db, "device_type"),
        recovery_statuses=active_items(db, "recovery_status"),
        recovery_reasons=active_items(db, "recovery_reason"),
        maintenance_contacts=ledger_service.contact_options(db),
        edit_device=edit_device,
        form_data=form_data or {},
        q=q,
        recovery_status_id=recovery_status_id or 0,
        page=page,
        total=total,
        last_page=last_page,
        error=error,
    )


@app.get("/devices")
def devices_page(
    request: Request,
    edit_id: int | None = None,
    q: str = "",
    recovery_status_id: int | None = None,
    page: int = 1,
    error: str = "",
    db: Session = Depends(get_db),
):
    auth.require_login(request)
    edit_device = db.get(NetworkDevice, edit_id) if edit_id else None
    return devices_template(
        request, db, edit_device=edit_device, q=q, recovery_status_id=recovery_status_id,
        page=page, error=error,
    )


def _parse_device_form(form) -> dict:
    return {
        "device_code": str(form.get("device_code", "")).strip(),
        "business_service_id": _form_int(form, "business_service_id"),
        "asset_class": str(form.get("asset_class", "")).strip(),
        "device_type": str(form.get("device_type", "")).strip(),
        "recovery_status": str(form.get("recovery_status", "")).strip(),
        "recovery_reason": str(form.get("recovery_reason", "")).strip(),
        "maintenance_contact_id": _form_int(form, "maintenance_contact_id"),
        "vendor_model": str(form.get("vendor_model", "")).strip(),
        "location": str(form.get("location", "")).strip(),
        "asset_value": _form_float(form, "asset_value"),
    }


def _apply_device_form(device: NetworkDevice, data: dict, db: Session) -> str | None:
    error = ledger_service.validate_device_code(db, data["device_code"], exclude_id=device.id)
    if error:
        return error
    if data["business_service_id"] is None:
        return "必须选择所属业务。"
    device.device_code = data["device_code"]
    device.business_service_id = data["business_service_id"]
    device.asset_class_item_id = _dictionary_value_id(db, "asset_class", data["asset_class"])
    device.device_type_item_id = _dictionary_value_id(db, "device_type", data["device_type"])
    device.recovery_status_item_id = _dictionary_value_id(db, "recovery_status", data["recovery_status"])
    device.recovery_reason_item_id = _dictionary_value_id(db, "recovery_reason", data["recovery_reason"])
    device.maintenance_contact_id = data["maintenance_contact_id"]
    device.vendor_model = data["vendor_model"]
    device.location = data["location"]
    device.asset_value = data["asset_value"]
    device.version = (device.version or 0) + 1
    return None


@app.post("/devices")
async def device_create(request: Request, db: Session = Depends(get_db)):
    auth.require_login(request)
    form = await request.form()
    try:
        data = _parse_device_form(form)
        device = NetworkDevice(device_code="", business_service_id=data["business_service_id"] or 0)
        error = _apply_device_form(device, data, db)
    except ValueError as exc:
        return redirect_to(f"/devices?error={quote(str(exc))}")
    if error:
        return redirect_to(f"/devices?error={quote(error)}")
    db.add(device)
    db.commit()
    return redirect_to("/devices")


@app.post("/devices/{device_id}/edit")
async def device_update(device_id: int, request: Request, db: Session = Depends(get_db)):
    auth.require_login(request)
    device = get_or_404(db, NetworkDevice, device_id)
    form = await request.form()
    try:
        data = _parse_device_form(form)
        error = _apply_device_form(device, data, db)
    except ValueError as exc:
        return redirect_to(f"/devices?error={quote(str(exc))}")
    if error:
        return redirect_to(f"/devices?error={quote(error)}")
    db.commit()
    return redirect_to("/devices")


@app.post("/devices/{device_id}/delete")
def device_delete(device_id: int, request: Request, db: Session = Depends(get_db)):
    auth.require_login(request)
    device = get_or_404(db, NetworkDevice, device_id)
    device.is_active = False
    db.commit()
    return redirect_to("/devices")


# ---------------------------------------------------------------- 审核中心


@app.get("/reviews")
def reviews_page(request: Request, status_filter: str = "", db: Session = Depends(get_db)):
    auth.require_login(request)
    query = select(ChangeSet).order_by(ChangeSet.created_at.desc()).limit(200)
    allowed_domains = [domain for domain in ("business", "network", "callback", "template", "system") if auth.has_permission(db, request, "read", domain)]
    if allowed_domains:
        query = query.where(ChangeSet.domain.in_(allowed_domains))
    else:
        query = query.where(False)
    if status_filter:
        query = query.where(ChangeSet.status == status_filter)
    change_sets = db.scalars(query).all()
    status_counts = {
        name: db.scalar(select(func.count(ChangeSet.id)).where(ChangeSet.status == name)) or 0
        for name in ["draft", "submitted", "returned", "approved", "rejected", "applied", "cancelled"]
    }
    return render(
        request,
        "reviews.html",
        db,
        change_sets=change_sets,
        status_filter=status_filter,
        status_counts=status_counts,
    )


# ---------------------------------------------------------------- 维系与回收工作台


def due_work_template(
    request: Request,
    db: Session,
    error: str = "",
    notice: str = "",
    status_code: int = 200,
):
    timezone_name = settings.default_timezone
    return render(
        request,
        "due_work.html",
        db,
        status_code=status_code,
        lead_days=change_request_service.due_renewal_lead_days(db),
        due_rows=change_request_service.list_due_renewal_rows(db, timezone_name=timezone_name),
        device_rows=change_request_service.list_recycle_device_rows(db, timezone_name=timezone_name),
        min_date=datetime.now(plan_service.get_zone(timezone_name)).date().isoformat(),
        error=error,
        notice=notice,
    )


@app.get("/due-work")
def due_work_page(
    request: Request,
    error: str = "",
    notice: str = "",
    db: Session = Depends(get_db),
):
    auth.require_login(request)
    return due_work_template(request, db, error=error, notice=notice)


def _submit_business_request(
    service_id: int,
    request: Request,
    db: Session,
    updates: dict,
    reason: str,
) -> RedirectResponse:
    """续签 / 退网共用提交路径：校验、落库、审计，错误回显工作台。"""

    service = get_or_404(db, BusinessService, service_id)
    try:
        change_set = change_request_service.submit_business_update(
            db, service, updates, reason, _user_id(request)
        )
        ledger_service.log_action(
            db, "submit_review", _user_id(request), "change_set", change_set.id,
            json.dumps({"domain": "business", "business_service_id": service.id, "updates": updates}, ensure_ascii=False),
            _client_ip(request),
        )
        db.commit()
    except ValueError as exc:
        db.rollback()
        return due_work_template(request, db, error=str(exc), status_code=400)
    return redirect_to(f"/reviews/{change_set.id}")


@app.post("/due-work/business/{service_id}/renew")
async def due_work_renew(service_id: int, request: Request, db: Session = Depends(get_db)):
    auth.require_permission(db, request, "submit", "business")
    form = await request.form()
    updates = {"agreement_expires_at": str(form.get("agreement_expires_at", "")).strip()}
    reason = str(form.get("reason", "")).strip()
    return _submit_business_request(service_id, request, db, updates, reason)


@app.post("/due-work/business/{service_id}/retire")
async def due_work_retire(service_id: int, request: Request, db: Session = Depends(get_db)):
    auth.require_permission(db, request, "submit", "business")
    form = await request.form()
    updates = {"service_status": change_request_service.RETIRE_STATUS_LABEL}
    reason = str(form.get("reason", "")).strip()
    return _submit_business_request(service_id, request, db, updates, reason)


@app.post("/due-work/device/{device_id}/recover")
async def due_work_recover(device_id: int, request: Request, db: Session = Depends(get_db)):
    auth.require_permission(db, request, "submit", "network")
    device = get_or_404(db, NetworkDevice, device_id)
    form = await request.form()
    reason = str(form.get("reason", "")).strip()
    try:
        change_set = change_request_service.submit_device_recovery(
            db, device, reason, _user_id(request)
        )
        ledger_service.log_action(
            db, "submit_review", _user_id(request), "change_set", change_set.id,
            json.dumps({"domain": "network", "operation": "recover", "device_id": device.id}, ensure_ascii=False),
            _client_ip(request),
        )
        db.commit()
    except ValueError as exc:
        db.rollback()
        return due_work_template(request, db, error=str(exc), status_code=400)
    return redirect_to(f"/reviews/{change_set.id}")


# ---------------------------------------------------------------- 缺项工作台


@app.get("/missing")
def missing_page(request: Request, db: Session = Depends(get_db)):
    auth.require_login(request)
    business_rows = ledger_service.business_missing_fields(db)
    device_rows = ledger_service.device_missing_fields(db)
    recent_batches = db.scalars(
        select(ImportBatch)
        .where(ImportBatch.missing_count > 0)
        .order_by(ImportBatch.created_at.desc())
        .limit(10)
    ).all()
    return render(
        request,
        "missing.html",
        db,
        business_rows=business_rows,
        device_rows=device_rows,
        recent_batches=recent_batches,
    )


# ---------------------------------------------------------------- 导入导出


def imports_template(
    request: Request,
    db: Session,
    error: str = "",
    status_code: int = 200,
):
    batches = db.scalars(select(ImportBatch).order_by(ImportBatch.created_at.desc()).limit(100)).all()
    return render(
        request,
        "imports.html",
        db,
        status_code=status_code,
        batches=batches,
        error=error,
    )


@app.get("/imports")
def imports_page(request: Request, db: Session = Depends(get_db)):
    auth.require_permission(db, request, "import", "system")
    return imports_template(request, db)


@app.post("/imports/upload")
async def import_upload(request: Request, file: UploadFile, db: Session = Depends(get_db)):
    auth.require_permission(db, request, "import", "system")
    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        return imports_template(request, db, "只支持 .xlsx 文件。", status_code=400)
    content = await file.read()
    if not content:
        return imports_template(request, db, "文件为空。", status_code=400)

    batch = ImportBatch(
        file_name=Path(filename).name,
        file_hash=hashlib.sha256(content).hexdigest(),
        source_type="business_ledger",
        status="validating",
        created_by_user_id=_user_id(request),
    )
    db.add(batch)
    db.flush()

    import_dir = BASE_DIR / "data" / "imports"
    stored = import_dir / f"{batch.id:05d}_{uuid.uuid4().hex[:8]}_{Path(filename).name}"
    stored.write_bytes(content)
    try:
        import_service.import_ledger_workbook(db, batch, stored)
    except Exception as exc:  # noqa: BLE001 - 解析失败要给用户可见错误
        db.rollback()
        return imports_template(request, db, f"导入解析失败：{exc}", status_code=400)
    ledger_service.log_action(
        db, "import", _user_id(request), "import_batch", batch.id,
        json.dumps({"file": batch.file_name, "rows": batch.total_rows}, ensure_ascii=False),
        _client_ip(request),
    )
    db.commit()
    return redirect_to(f"/imports/{batch.id}")


@app.get("/imports/{batch_id}")
def import_detail(
    batch_id: int,
    request: Request,
    status_filter: str = "",
    row_query: str = "",
    edit_row_id: int | None = None,
    error: str = "",
    status_code: int = 200,
    db: Session = Depends(get_db),
):
    auth.require_permission(db, request, "import", "system")
    batch = get_or_404(db, ImportBatch, batch_id)
    query = select(StagingRow).where(StagingRow.batch_id == batch.id).order_by(
        StagingRow.row_number.asc()
    )
    if status_filter:
        query = query.where(StagingRow.status == status_filter)
    all_rows = db.scalars(query).all()
    if row_query.strip():
        needle = row_query.strip().casefold()
        rows = []
        for row in all_rows:
            haystack = " ".join(
                [str(row.row_number), row.error_messages or "", row.mapped_json or "", row.raw_json or ""]
            ).casefold()
            if needle in haystack:
                rows.append(row)
        rows = rows[:500]
    else:
        rows = all_rows[:500]
    status_counts = {
        name: db.scalar(
            select(func.count(StagingRow.id)).where(
                StagingRow.batch_id == batch.id, StagingRow.status == name
            )
        ) or 0
        for name in ["valid", "missing", "error", "duplicate"]
    }
    edit_row = db.get(StagingRow, edit_row_id) if edit_row_id else None
    if edit_row is not None and edit_row.batch_id != batch.id:
        edit_row = None
    edit_raw = {}
    if edit_row is not None:
        edit_payload = json.loads(edit_row.mapped_json or "{}")
        edit_raw = edit_payload.get("manual_raw", json.loads(edit_row.raw_json or "{}"))
        for field in ("入网时间", "协议到期时间"):
            edit_raw[field] = html_date_value(edit_raw.get(field, ""))
    customers = db.scalars(select(Customer).where(Customer.is_active.is_(True)).order_by(Customer.name.asc())).all()
    channel_options = [value for value in db.scalars(
        select(BusinessService.channel_name).where(BusinessService.channel_name != "").distinct().order_by(BusinessService.channel_name)
    ).all() if value]
    device_codes = [value for value in db.scalars(
        select(NetworkDevice.device_code).where(NetworkDevice.device_code != "").distinct().order_by(NetworkDevice.device_code)
    ).all() if value]
    vendor_models = [value for value in db.scalars(
        select(NetworkDevice.vendor_model).where(NetworkDevice.vendor_model != "").distinct().order_by(NetworkDevice.vendor_model)
    ).all() if value]
    device_locations = [value for value in db.scalars(
        select(NetworkDevice.location).where(NetworkDevice.location != "").distinct().order_by(NetworkDevice.location)
    ).all() if value]
    asset_values = [str(value) for value in db.scalars(
        select(NetworkDevice.asset_value).where(NetworkDevice.asset_value.is_not(None)).distinct().order_by(NetworkDevice.asset_value)
    ).all() if value is not None]
    directory_contacts = db.scalars(
        select(Contact).where(Contact.is_active.is_(True)).order_by(Contact.name.asc(), Contact.id.asc())
    ).all()
    return render(
        request,
        "import_detail.html",
        db,
        status_code=status_code,
        batch=batch,
        rows=rows,
        status_filter=status_filter,
        row_query=row_query,
        status_counts=status_counts,
        edit_row=edit_row,
        edit_raw=edit_raw,
        import_columns=import_service.LEDGER_COLUMNS + import_service.TECHNICAL_COLUMNS,
        business_columns=import_service.BUSINESS_COLUMNS,
        device_columns=import_service.DEVICE_COLUMNS,
        customers=customers,
        counties=active_items(db, "county"),
        grids=active_items(db, "grid"),
        service_statuses=active_items(db, "service_status"),
        business_types=active_items(db, "business_type"),
        asset_classes=active_items(db, "asset_class"),
        device_types=active_items(db, "device_type"),
        recovery_statuses=active_items(db, "recovery_status"),
        recovery_reasons=active_items(db, "recovery_reason"),
        contact_options=directory_contacts,
        channel_options=channel_options,
        device_codes=device_codes,
        vendor_models=vendor_models,
        device_locations=device_locations,
        asset_values=asset_values,
        error=error,
    )


@app.post("/imports/{batch_id}/submit")
def import_submit(batch_id: int, request: Request, db: Session = Depends(get_db)):
    auth.require_permission(db, request, "import", "system")
    auth.require_permission(db, request, "submit", "business")
    auth.require_permission(db, request, "submit", "network")
    batch = get_or_404(db, ImportBatch, batch_id)
    try:
        change_sets = review_service.build_import_change_sets(db, batch, _user_id(request))
        ledger_service.log_action(
            db,
            "submit_review",
            _user_id(request),
            "import_batch", batch.id,
            json.dumps({"change_set_ids": [item.id for item in change_sets]}, ensure_ascii=False),
            _client_ip(request),
        )
        db.commit()
    except review_service.ChangeApplicationError as exc:
        db.rollback()
        return import_detail(batch_id, request, error=str(exc), status_code=409, db=db)
    return redirect_to(f"/reviews/{change_sets[0].id}")


@app.post("/imports/{batch_id}/rows/{row_id}/correct")
async def import_row_correct(batch_id: int, row_id: int, request: Request, db: Session = Depends(get_db)):
    auth.require_permission(db, request, "import", "system")
    batch = get_or_404(db, ImportBatch, batch_id)
    if batch.status != "ready":
        return import_detail(batch_id, request, error="只有待提交批次可以更正暂存行。", db=db, status_code=400)
    form = await request.form()
    try:
        import_service.revalidate_staging_row(
            db, batch, row_id,
            {column: str(form.get(column, "")) for column in import_service.LEDGER_COLUMNS + import_service.TECHNICAL_COLUMNS},
        )
        db.commit()
    except ValueError as exc:
        db.rollback()
        return import_detail(batch_id, request, error=str(exc), db=db, status_code=400)
    return redirect_to(f"/imports/{batch.id}?edit_row_id={row_id}")


def _review_template(
    request: Request,
    db: Session,
    change_set: ChangeSet,
    error: str = "",
    notice: str = "",
    status_code: int = 200,
):
    return render(
        request,
        "review_detail.html",
        db,
        status_code=status_code,
        change_set=change_set,
        related_change_sets=(db.scalars(select(ChangeSet).where(ChangeSet.import_batch_id == change_set.import_batch_id, ChangeSet.id != change_set.id).order_by(ChangeSet.domain)).all() if change_set.import_batch_id else []),
        previews=[review_service.preview_change_item(db, item) for item in change_set.items],
        error=error,
        notice=notice,
        is_self_submission=_is_self_submission(request, change_set),
        can_review=change_set.status == "submitted"
        and auth.has_permission(db, request, "review", change_set.domain)
        and (not _is_self_submission(request, change_set) or auth.is_system_admin(db, request)),
        can_apply=change_set.status == "approved"
        and auth.has_permission(db, request, "apply", change_set.domain),
)


@app.get("/reviews/{change_set_id}")
def review_detail(change_set_id: int, request: Request, db: Session = Depends(get_db)):
    change_set = get_or_404(db, ChangeSet, change_set_id)
    auth.require_permission(db, request, "read", change_set.domain)
    return _review_template(request, db, change_set)


@app.post("/reviews/{change_set_id}/decision")
async def review_decision(change_set_id: int, request: Request, db: Session = Depends(get_db)):
    change_set = get_or_404(db, ChangeSet, change_set_id)
    auth.require_permission(db, request, "review", change_set.domain)
    form = await request.form()
    decision = str(form.get("decision", ""))
    reason = str(form.get("reason", "")).strip()
    user_id = _user_id(request)
    if change_set.status != "submitted":
        return _review_template(request, db, change_set, "该申请当前不可审核。", status_code=400)
    self_review = _is_self_submission(request, change_set)
    if self_review and not auth.is_system_admin(db, request):
        return _review_template(request, db, change_set, "提交人不能审核自己的申请。", status_code=403)
    if self_review and str(form.get("self_review_confirmed", "")) != "1":
        return _review_template(request, db, change_set, "自提交、自审核必须勾选二次确认。", status_code=400)
    if decision not in {"approved", "returned", "rejected"}:
        return _review_template(request, db, change_set, "无效的审核操作。", status_code=400)
    if decision == "returned" and not reason:
        return _review_template(request, db, change_set, "退回时必须填写原因。", status_code=400)
    change_set.status = decision
    change_set.reviewed_by_user_id = user_id
    change_set.reviewed_at = utcnow()
    if reason:
        change_set.reason = reason
    for item in change_set.items:
        payload = json.loads(item.patch_json)
        batch_id = payload.get("batch_id")
        batch = db.get(ImportBatch, batch_id) if batch_id else None
        if batch is not None:
            batch.reviewed_by_user_id = user_id
            batch.status = "rejected" if decision == "rejected" else "ready" if decision == "returned" else "reviewing"
    ledger_service.log_action(
        db, f"review_{decision}", user_id, "change_set", change_set.id,
        json.dumps({"reason": reason, "self_review": self_review}, ensure_ascii=False), _client_ip(request)
    )
    db.commit()
    return redirect_to(f"/reviews/{change_set.id}")


@app.post("/reviews/{change_set_id}/apply")
def review_apply(change_set_id: int, request: Request, db: Session = Depends(get_db)):
    change_set = get_or_404(db, ChangeSet, change_set_id)
    auth.require_permission(db, request, "apply", change_set.domain)
    try:
        applied = review_service.apply_change_set(db, change_set, _user_id(request))
        ledger_service.log_action(
            db, "apply_change_set", _user_id(request), "change_set", change_set.id,
            json.dumps({"items": applied}, ensure_ascii=False), _client_ip(request)
        )
        db.commit()
    except review_service.ChangeApplicationError as exc:
        db.rollback()
        change_set = get_or_404(db, ChangeSet, change_set_id)
        return _review_template(request, db, change_set, str(exc), status_code=409)
    return redirect_to(f"/reviews/{change_set.id}")


def build_ledger_export_workbook(db: Session):
    """Build one Chinese data sheet plus a human-readable instruction sheet."""

    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    workbook = Workbook()
    ledger_sheet = workbook.active
    ledger_sheet.title = "业务设备台账"
    ledger_sheet.append(["业务信息"] + [None] * 14 + ["设备信息"] + [None] * 7 + ["系统校验字段（请勿修改）"] + [None] * 3)
    ledger_sheet.append(import_service.LEDGER_COLUMNS + import_service.TECHNICAL_COLUMNS)
    services = db.scalars(select(BusinessService).where(BusinessService.is_active.is_(True)).order_by(BusinessService.service_number)).all()
    for service in services:
        devices = [device for device in service.devices if device.is_active] or [None]
        for device in devices:
            ledger_sheet.append([
                service.service_number, service.customer.name, service.county_item.label if service.county_item else "", service.grid_item.label if service.grid_item else "",
                service.service_status_item.label if service.service_status_item else "", format_date(service.accessed_at), format_date(service.agreement_expires_at), service.business_type_item.label if service.business_type_item else "", service.channel_name,
                "", "", "", "", "", "",
                device.asset_class_item.label if device and device.asset_class_item else "", device.device_code if device else "", device.asset_value if device else "", device.device_type_item.label if device and device.device_type_item else "", device.vendor_model if device else "", device.location if device else "", device.recovery_status_item.label if device and device.recovery_status_item else "", device.recovery_reason_item.label if device and device.recovery_reason_item else "",
                service.id, service.version, device.id if device else "", device.version if device else "",
            ])
    ledger_sheet.freeze_panes = "A3"
    ledger_sheet.auto_filter.ref = f"A2:{ledger_sheet.cell(2, len(import_service.LEDGER_COLUMNS + import_service.TECHNICAL_COLUMNS)).column_letter}{ledger_sheet.max_row}"
    guidance = {field: (requirement, example) for field, requirement, example in import_service.LEDGER_FIELD_GUIDANCE}
    for cell in ledger_sheet[2]:
        if cell.value in guidance:
            requirement, example = guidance[cell.value]
            cell.comment = Comment(f"{requirement}\n示例：{example}", "中国联通 IT 运维系统")
    for cell in ledger_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="3F4A46")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for cell in ledger_sheet[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8ECE9")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    technical_start = len(import_service.LEDGER_COLUMNS) + 1
    for row in ledger_sheet.iter_rows(min_row=2, min_col=technical_start, max_col=ledger_sheet.max_column):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="F1F3F2")

    validation_options = {
        "E": ["正常开机", "主动退网(申请拆机)"],
        "H": ["数据及网元业务", "宽带业务"],
        "P": ["资产类", "成本类"],
        "V": ["已回收", "未回收"],
        "W": ["在用", "遗失", "纠纷", "其他"],
    }
    for column_letter, options in validation_options.items():
        validation = DataValidation(
            type="list",
            formula1=f'"{",".join(options)}"',
            allow_blank=True,
            errorTitle="填写值不符合要求",
            error="请从下拉列表中选择允许值。",
            promptTitle="字段填写要求",
            prompt="请使用下拉列表中的标准值。",
        )
        validation.errorStyle = "stop"
        validation.showErrorMessage = True
        validation.showInputMessage = True
        ledger_sheet.add_data_validation(validation)
        validation.add(f"{column_letter}3:{column_letter}10000")

    instruction_sheet = workbook.create_sheet("填写说明")
    instruction_sheet.merge_cells("A1:C1")
    instruction_sheet["A1"] = "业务设备台账填写说明"
    instruction_sheet["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    instruction_sheet["A1"].fill = PatternFill("solid", fgColor="3F4A46")
    instruction_sheet["A1"].alignment = Alignment(horizontal="center")
    instruction_sheet.append(["字段", "填写要求", "示例"])
    for field, requirement, example in import_service.LEDGER_FIELD_GUIDANCE:
        instruction_sheet.append([field, requirement, example])
    instruction_sheet.freeze_panes = "A3"
    instruction_sheet.column_dimensions["A"].width = 28
    instruction_sheet.column_dimensions["B"].width = 88
    instruction_sheet.column_dimensions["C"].width = 28
    for cell in instruction_sheet[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8ECE9")
    for row in instruction_sheet.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column in ledger_sheet.columns:
        ledger_sheet.column_dimensions[column[0].column_letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 28)
    return workbook


@app.get("/exports/ledger.xlsx")
def export_ledger(request: Request, db: Session = Depends(get_db)):
    auth.require_permission(db, request, "export", "system")
    workbook = build_ledger_export_workbook(db)
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    ledger_service.log_action(
        db, "export", _user_id(request), "business_ledger", detail=json.dumps({"format": "flat_zh_xlsx"}, ensure_ascii=False), ip_address=_client_ip(request)
    )
    db.commit()
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=unicom-it-ledger-export.xlsx"},
    )


# ---------------------------------------------------------------- 备份与灾备


def _backups_template(
    request: Request,
    db: Session,
    error: str = "",
    success: str = "",
    status_code: int = 200,
):
    return render(
        request,
        "admin_backups.html",
        db,
        status_code=status_code,
        backups=backup_service.list_backups(),
        backup_status=backup_service.status(),
        error=error,
        success=success,
    )


@app.get("/admin/backups")
def admin_backups_page(request: Request, db: Session = Depends(get_db)):
    auth.require_permission(db, request, "manage_config", "system")
    return _backups_template(request, db)


@app.post("/admin/backups")
def admin_backup_create(request: Request, db: Session = Depends(get_db)):
    auth.require_permission(db, request, "manage_config", "system")
    try:
        info = backup_service.create_backup()
    except Exception as exc:  # noqa: BLE001 - display operational failure to admin
        ledger_service.log_action(
            db, "backup_failed", _user_id(request), "backup",
            detail=json.dumps({"error": str(exc)[:500]}, ensure_ascii=False),
            ip_address=_client_ip(request),
        )
        db.commit()
        return _backups_template(request, db, error=f"备份失败：{exc}", status_code=500)
    ledger_service.log_action(
        db, "backup_created", _user_id(request), "backup",
        detail=json.dumps({"filename": info.filename, "remote_uploaded": info.remote_uploaded}, ensure_ascii=False),
        ip_address=_client_ip(request),
    )
    db.commit()
    remote_note = "，已上传 WebDAV" if info.remote_uploaded else ""
    return _backups_template(request, db, success=f"备份已创建：{info.filename}{remote_note}")


@app.get("/admin/backups/{filename}/download")
def admin_backup_download(filename: str, request: Request, db: Session = Depends(get_db)):
    auth.require_permission(db, request, "manage_config", "system")
    path = backup_service.backup_directory / filename
    if (
        Path(filename).name != filename
        or not filename.startswith("callback-backup-v")
        or path.suffix != ".zip"
        or not path.is_file()
    ):
        raise HTTPException(status_code=404)
    ledger_service.log_action(
        db, "backup_download", _user_id(request), "backup",
        detail=json.dumps({"filename": filename}, ensure_ascii=False),
        ip_address=_client_ip(request),
    )
    db.commit()
    return FileResponse(path, media_type="application/zip", filename=filename)


# ---------------------------------------------------------------- 系统管理


@app.get("/admin/users")
def admin_users_page(request: Request, edit_id: int | None = None, db: Session = Depends(get_db)):
    auth.require_permission(db, request, "manage_users", "system")
    return _users_page(request, db, edit_user=db.get(User, edit_id) if edit_id else None)


def _users_page(
    request: Request,
    db: Session,
    error: str | None = None,
    edit_user: User | None = None,
    status_code: int = 200,
):
    """/admin/users 页面渲染：统一错误回显所需的列表与角色上下文。"""
    users = db.scalars(select(User).order_by(User.created_at.asc())).all()
    roles = db.scalars(select(Role).order_by(Role.id.asc())).all()
    user_role_map = {
        user.id: db.scalars(
            select(UserRole.role_id).where(UserRole.user_id == user.id)
        ).all()
        for user in users
    }
    return render(
        request,
        "admin_users.html",
        db,
        users=users,
        roles=roles,
        user_role_map=user_role_map,
        role_name_map={user.id: role_names(db, user) for user in users},
        edit_user=edit_user,
        error=error,
        status_code=status_code,
    )


@app.post("/admin/users")
async def admin_user_create(request: Request, db: Session = Depends(get_db)):
    auth.require_permission(db, request, "manage_users", "system")
    form = await request.form()
    username = str(form.get("username", "")).strip()
    display_name = str(form.get("display_name", "")).strip()
    real_name = str(form.get("real_name", "")).strip()
    phone = str(form.get("phone", "")).strip()
    password = str(form.get("password", ""))
    role_ids = [int(v) for v in form.getlist("role_ids")]
    enabled = str(form.get("enabled", "")) == "on"
    if not username or not password:
        return _users_page(request, db, error="用户名和初始密码必填。", status_code=400)
    error = validate_user_profile(real_name, phone)
    if error:
        return _users_page(request, db, error=error, status_code=400)
    user = User(
        username=username,
        display_name=display_name,
        real_name=real_name,
        phone=phone,
        password_hash=hash_password(password),
        is_enabled=enabled,
        is_superadmin=False,
    )
    db.add(user)
    try:
        db.flush()
        set_user_roles(db, user, role_ids)
        db.commit()
    except IntegrityError:
        db.rollback()
        return _users_page(request, db, error=f"用户名 {username} 已存在。", status_code=400)
    ledger_service.log_action(
        db, "manage_users", _user_id(request), "user", user.id,
        json.dumps({"operation": "create", "username": username}, ensure_ascii=False),
        _client_ip(request),
    )
    db.commit()
    return redirect_to("/admin/users")


@app.post("/admin/users/{user_id}/edit")
async def admin_user_update(user_id: int, request: Request, db: Session = Depends(get_db)):
    auth.require_permission(db, request, "manage_users", "system")
    user = get_or_404(db, User, user_id)
    form = await request.form()
    real_name = str(form.get("real_name", "")).strip()
    phone = str(form.get("phone", "")).strip()
    error = validate_user_profile(real_name, phone, is_superadmin=user.is_superadmin)
    if error:
        return _users_page(request, db, error=error, edit_user=user, status_code=400)
    user.real_name = real_name
    user.phone = phone
    user.display_name = str(form.get("display_name", "")).strip()
    user.is_enabled = str(form.get("enabled", "")) == "on"
    role_ids = [int(v) for v in form.getlist("role_ids")]
    set_user_roles(db, user, role_ids)
    db.commit()
    ledger_service.log_action(
        db, "manage_users", _user_id(request), "user", user.id,
        json.dumps({"operation": "update", "username": user.username}, ensure_ascii=False),
        _client_ip(request),
    )
    db.commit()
    return redirect_to("/admin/users")


@app.post("/admin/users/{user_id}/password")
async def admin_user_password(user_id: int, request: Request, db: Session = Depends(get_db)):
    auth.require_permission(db, request, "manage_users", "system")
    user = get_or_404(db, User, user_id)
    form = await request.form()
    password = str(form.get("password", ""))
    if len(password) < 8:
        return _users_page(request, db, error="密码至少 8 位。", edit_user=user, status_code=400)
    user.password_hash = hash_password(password)
    # 管理员重置后视为发放新初始密码，被重置账号下次登录须改密。
    user.force_password_change = True
    db.commit()
    return redirect_to("/admin/users")


@app.get("/admin/roles")
def admin_roles_page(request: Request, db: Session = Depends(get_db)):
    auth.require_permission(db, request, "manage_users", "system")
    roles = db.scalars(select(Role).order_by(Role.id.asc())).all()
    permissions = db.scalars(
        select(RolePermission).order_by(RolePermission.role_id.asc(), RolePermission.domain.asc())
    ).all()
    from .models import Permission

    permission_map = {p.id: p for p in db.scalars(select(Permission)).all()}
    return render(
        request,
        "admin_roles.html",
        db,
        roles=roles,
        permissions=permissions,
        permission_map=permission_map,
    )


# ---------------------------------------------------------------- 系统监控


SYSTEM_TABS = {"status", "logs", "settings"}

# 监控页展示的 .env 配置（密钥类配置一律不展示）。
ENV_SETTINGS = [
    ("MODEM_PORT", lambda s: s.modem_port, "A7670E 串口设备路径"),
    ("MODEM_BAUD", lambda s: s.modem_baud, "串口波特率"),
    ("AUDIO_DEVICE", lambda s: s.audio_device, "ffmpeg 播放设备（ALSA）"),
    ("CALL_CONNECT_TIMEOUT_SECONDS", lambda s: s.call_connect_timeout_seconds, "接通等待超时兜底（秒）"),
    ("MIN_CONNECTED_SECONDS", lambda s: s.min_connected_seconds, "接通后有效时长阈值（秒）"),
    ("RETRY_DELAY_SECONDS", lambda s: s.retry_delay_seconds, "自动重试延迟（秒）"),
    ("MAX_CALL_ATTEMPTS", lambda s: s.max_call_attempts, "最大尝试次数（含首次）"),
    ("TTS_PROVIDER", lambda s: s.tts_provider, "TTS 提供商（none 表示离线）"),
    ("DEFAULT_TIMEZONE", lambda s: s.default_timezone, "计划默认时区"),
    ("CALL_WORKER_ENABLED", lambda s: "是" if s.call_worker_enabled else "否", "外呼 Worker 自动启动硬开关"),
    ("WORKER_POLL_SECONDS", lambda s: s.worker_poll_seconds, "Worker 空闲轮询间隔（秒）"),
    ("BACKUP_ENABLED", lambda s: "是" if s.backup_enabled else "否", "后台定期备份开关"),
    ("BACKUP_INTERVAL_HOURS", lambda s: s.backup_interval_hours, "备份间隔（小时）"),
    ("BACKUP_RETENTION_DAYS", lambda s: s.backup_retention_days, "本地备份保留天数"),
    ("BACKUP_DIR", lambda s: s.backup_dir, "本地备份目录"),
    ("BACKUP_MAX_RETRIES", lambda s: s.backup_max_retries, "远端上传失败重试次数"),
    ("BACKUP_WEBDAV_URL", lambda s: "已配置" if s.backup_webdav_url else "未配置", "WebDAV 目标（凭据不展示）"),
]


def _day_start_utc() -> datetime:
    zone = plan_service.get_zone(settings.default_timezone)
    today = datetime.now(zone).date()
    return datetime.combine(today, time.min, tzinfo=zone).astimezone(timezone.utc)


def _schema_status(db: Session) -> dict:
    current = db.execute(text("SELECT version_num FROM alembic_version")).scalar()
    head = get_schema_head()
    return {"current": current or "", "head": head, "ok": current == head}


@app.get("/admin/system")
def admin_system_page(
    request: Request,
    tab: str = "status",
    event_type: str = "",
    db: Session = Depends(get_db),
):
    tab = tab if tab in SYSTEM_TABS else "status"
    auth.require_permission(
        db, request, "read" if tab == "logs" else "manage_config", "system"
    )

    if tab == "status":
        scripts = db.scalars(select(Script)).all()
        scripts_with_wav = sum(1 for s in scripts if s.wav_path and Path(s.wav_path).exists())
        worker_status = call_worker_service.status()
        queue_counts = {
            name: db.scalar(select(func.count(CallTask.id)).where(CallTask.status == name)) or 0
            for name in ["queued", "dialing"]
        }
        today_records = db.scalar(
            select(func.count(CallRecord.id)).where(CallRecord.created_at >= _day_start_utc())
        ) or 0
        return render(
            request,
            "admin_system.html",
            db,
            tab=tab,
            scheduler_status=scheduler_service.status(),
            scheduler_enabled=is_scheduler_enabled(db),
            worker_status=worker_status,
            worker_enabled=is_worker_enabled(db),
            worker_gate=settings.call_worker_enabled,
            modem_status=modem_availability(settings, worker_status),
            queue_counts=queue_counts,
            scripts_total=len(scripts),
            scripts_with_wav=scripts_with_wav,
            today_records=today_records,
            schema=_schema_status(db),
            backup_status=backup_service.status(),
        )

    if tab == "logs":
        event_types = db.scalars(
            select(CallEvent.event_type).distinct().order_by(CallEvent.event_type)
        ).all()
        events_query = select(CallEvent).order_by(CallEvent.created_at.desc())
        if event_type:
            events_query = events_query.where(CallEvent.event_type == event_type)
        recent_events = db.scalars(events_query.limit(100)).all()
        recent_audit = db.scalars(
            select(AuditLog).order_by(AuditLog.created_at.desc()).limit(50)
        ).all()
        return render(
            request,
            "admin_system.html",
            db,
            tab=tab,
            event_types=event_types,
            event_type_filter=event_type,
            recent_events=recent_events,
            recent_audit=recent_audit,
        )

    return render(
        request,
        "admin_system.html",
        db,
        tab=tab,
        env_settings=[(key, getter(settings), desc) for key, getter, desc in ENV_SETTINGS],
        runtime_settings=[
            ("scheduler_enabled", "1" if is_scheduler_enabled(db) else "0", "定时调度器开关"),
            (CALL_WORKER_ENABLED_KEY, "1" if is_worker_enabled(db) else "0", "外呼 Call Worker 开关"),
        ],
    )


# ---------------------------------------------------------------- 话术


def scripts_template(
    request: Request,
    db: Session,
    error: str = "",
    notice: str = "",
    form_data: dict[str, str] | None = None,
    edit_script: Script | None = None,
    status_code: int = 200,
):
    scripts = db.scalars(select(Script).order_by(Script.created_at.desc())).all()
    return render(
        request,
        "scripts.html",
        db,
        status_code=status_code,
        scripts=scripts,
        edit_script=edit_script,
        form_data=form_data or {},
        error=error,
        notice=notice,
    )


@app.get("/scripts")
def scripts_page(
    request: Request,
    edit_id: int | None = None,
    notice: str = "",
    db: Session = Depends(get_db),
):
    auth.require_login(request)
    edit_script = db.get(Script, edit_id) if edit_id else None
    return scripts_template(request, db, edit_script=edit_script, notice=notice)


@app.post("/scripts")
async def script_create(request: Request, db: Session = Depends(get_db)):
    auth.require_login(request)
    form = await request.form()
    title = str(form.get("title", "")).strip()
    body = str(form.get("body", "")).strip()
    wav_path = str(form.get("wav_path", "")).strip()
    form_data = {"title": title, "body": body, "wav_path": wav_path}
    if not title or not body:
        return scripts_template(request, db, "标题和话术内容必填。", form_data, status_code=400)
    # 手动指定 WAV 时视为已有可用音频；否则待生成。
    db.add(
        Script(
            title=title,
            body=body,
            wav_path=wav_path,
            tts_status="generated" if wav_path else "not_generated",
        )
    )
    db.commit()
    return redirect_to("/scripts")


@app.post("/scripts/{script_id}/edit")
async def script_update(script_id: int, request: Request, db: Session = Depends(get_db)):
    auth.require_login(request)
    script = get_or_404(db, Script, script_id)
    form = await request.form()
    title = str(form.get("title", "")).strip()
    body = str(form.get("body", "")).strip()
    wav_path = str(form.get("wav_path", "")).strip()
    if not title or not body:
        return scripts_template(request, db, "标题和话术内容必填。", edit_script=script, status_code=400)
    script.title = title
    script.body = body
    script.wav_path = wav_path
    if wav_path:
        script.tts_status = "generated"
        script.tts_error = ""
    db.commit()
    return redirect_to("/scripts")


@app.post("/scripts/{script_id}/delete")
def script_delete(script_id: int, request: Request, db: Session = Depends(get_db)):
    auth.require_login(request)
    script = get_or_404(db, Script, script_id)
    refs = script_referencing_counts(db, script)
    if any(refs.values()):
        parts = []
        for label, key in [
            ("扫描配置", "schedules"),
            ("回访计划", "plans"),
            ("外呼任务", "tasks"),
            ("通话记录", "records"),
        ]:
            if refs[key]:
                parts.append(f"{label} {refs[key]} 条")
        return scripts_template(
            request,
            db,
            f"该话术仍被引用（{'、'.join(parts)}），不能删除；可先更换引用的话术。",
            status_code=400,
        )
    try:
        db.delete(script)
        db.commit()
    except IntegrityError:
        db.rollback()
        return scripts_template(
            request,
            db,
            "该话术被其他数据引用，不能删除。",
            status_code=400,
        )
    return redirect_to("/scripts")


@app.post("/scripts/{script_id}/generate-audio")
def script_generate_audio(script_id: int, request: Request, db: Session = Depends(get_db)):
    auth.require_login(request)
    script = get_or_404(db, Script, script_id)
    message = generate_script_audio(db, script, settings)
    db.commit()
    return redirect_to(f"/scripts?notice={quote(message)}")


@app.get("/audio/{filename:path}")
def script_audio_file(filename: str, request: Request):
    """只读话术音频：仅服务 data/audio/ 下的 WAV/MP3，需登录，防路径穿越。"""
    auth.require_login(request)
    path = resolve_audio_file(filename)
    if path is None:
        raise HTTPException(status_code=404)
    media_type = "audio/mpeg" if path.suffix.lower() == ".mp3" else "audio/wav"
    return FileResponse(path, media_type=media_type)


# ---------------------------------------------------------------- 扫描通知配置


def scan_schedules_template(
    request: Request,
    db: Session,
    error: str = "",
    notice: str = "",
    form_data: dict[str, str] | None = None,
    edit_schedule: ScanSchedule | None = None,
    status_code: int = 200,
):
    schedules = db.scalars(
        select(ScanSchedule).order_by(ScanSchedule.created_at.desc())
    ).all()
    scripts = db.scalars(select(Script).order_by(Script.title.asc())).all()
    return render(
        request,
        "scan_schedules.html",
        db,
        status_code=status_code,
        schedules=schedules,
        scripts=scripts,
        edit_schedule=edit_schedule,
        form_data=form_data or {},
        error=error,
        notice=notice,
    )


def _parse_scan_schedule_form(db: Session, form) -> dict:
    name = str(form.get("name", "")).strip()
    if not name:
        raise ValueError("扫描配置名称不能为空。")
    scan_type = str(form.get("scan_type", "")).strip()
    if scan_type not in SCAN_TYPE_LABELS:
        raise ValueError("扫描类型无效。")
    script_id = _form_int(form, "script_id")
    if script_id is not None and db.get(Script, script_id) is None:
        raise ValueError("所选话术模板不存在。")
    cron_expr = str(form.get("cron_expr", "")).strip()
    timezone_name = str(form.get("timezone", "")).strip() or settings.default_timezone
    plan_service.validate_cron_expr(cron_expr, timezone_name)
    lead_days_raw = str(form.get("lead_days", "")).strip()
    if lead_days_raw:
        try:
            lead_days = int(lead_days_raw)
        except ValueError as exc:
            raise ValueError("提前天数必须是整数（例如 14）。") from exc
    else:
        lead_days = 14
    if lead_days < 0:
        raise ValueError("提前天数不能为负数。")
    enabled = str(form.get("enabled", "")) == "on"
    sms_enabled = str(form.get("sms_enabled", "")) == "on"
    return {
        "name": name,
        "scan_type": scan_type,
        "script_id": script_id,
        "cron_expr": cron_expr,
        "timezone": timezone_name,
        "lead_days": lead_days,
        "enabled": enabled,
        "sms_enabled": sms_enabled,
    }


@app.get("/scan-schedules")
def scan_schedules_page(
    request: Request,
    edit_id: int | None = None,
    notice: str = "",
    db: Session = Depends(get_db),
):
    auth.require_login(request)
    edit_schedule = db.get(ScanSchedule, edit_id) if edit_id else None
    return scan_schedules_template(request, db, edit_schedule=edit_schedule, notice=notice)


@app.post("/scan-schedules")
async def scan_schedule_create(request: Request, db: Session = Depends(get_db)):
    auth.require_login(request)
    form = await request.form()
    try:
        data = _parse_scan_schedule_form(db, form)
    except ValueError as exc:
        return scan_schedules_template(
            request, db, str(exc), form_data=dict(form), status_code=400
        )
    db.add(ScanSchedule(**data))
    db.commit()
    return redirect_to(f"/scan-schedules?notice={quote('扫描配置已创建。')}")


@app.post("/scan-schedules/{schedule_id}/edit")
async def scan_schedule_update(schedule_id: int, request: Request, db: Session = Depends(get_db)):
    auth.require_login(request)
    schedule = get_or_404(db, ScanSchedule, schedule_id)
    form = await request.form()
    try:
        data = _parse_scan_schedule_form(db, form)
    except ValueError as exc:
        return scan_schedules_template(
            request, db, str(exc), form_data=dict(form), edit_schedule=schedule, status_code=400
        )
    for key, value in data.items():
        setattr(schedule, key, value)
    db.commit()
    notice = f"扫描配置「{schedule.name}」已保存。"
    return redirect_to(f"/scan-schedules?edit_id={schedule.id}&notice={quote(notice)}")


@app.post("/scan-schedules/{schedule_id}/delete")
def scan_schedule_delete(schedule_id: int, request: Request, db: Session = Depends(get_db)):
    auth.require_login(request)
    schedule = get_or_404(db, ScanSchedule, schedule_id)
    # SQLite 默认不强制外键，显式检查引用，避免删除仍被任务引用的配置。
    task_refs = db.scalar(
        select(func.count(CallTask.id)).where(CallTask.scan_schedule_id == schedule.id)
    ) or 0
    if task_refs:
        return scan_schedules_template(
            request,
            db,
            f"该扫描配置已生成 {task_refs} 条通知任务，不能删除；如需停止扫描请先停用。",
            status_code=400,
        )
    try:
        db.delete(schedule)
        db.commit()
    except IntegrityError:
        db.rollback()
        return scan_schedules_template(
            request,
            db,
            "该扫描配置被其他数据引用，不能删除；如需停止扫描请先停用。",
            status_code=400,
        )
    return redirect_to("/scan-schedules")


@app.post("/scan-schedules/{schedule_id}/toggle")
def scan_schedule_toggle(schedule_id: int, request: Request, db: Session = Depends(get_db)):
    auth.require_login(request)
    schedule = get_or_404(db, ScanSchedule, schedule_id)
    schedule.enabled = not schedule.enabled
    db.commit()
    return redirect_to("/scan-schedules")


# ---------------------------------------------------------------- 短信通知

SMS_PAGE_LIMIT = 200


@app.get("/sms")
def sms_page(request: Request, db: Session = Depends(get_db)):
    auth.require_login(request)
    records = db.scalars(
        select(SmsNotification)
        .order_by(SmsNotification.id.desc())
        .limit(SMS_PAGE_LIMIT)
    ).all()
    return render(
        request,
        "sms.html",
        db,
        records=records,
        limit=SMS_PAGE_LIMIT,
    )


# ---------------------------------------------------------------- 通话记录


CALLS_PAGE_SIZE = 25
# 通话记录列表的候选状态筛选值（与 models.CALL_STATUSES 一致）。
CALL_STATUS_OPTIONS = sorted(CALL_STATUSES)


def _date_filter_bound(value: str, day_start: bool) -> datetime | None:
    """把 date 输入（YYYY-MM-DD）转成当日 00:00 / 23:59（计划时区），供区间筛选。"""

    value = value.strip()
    if not value:
        return None
    if len(value) == 10 and value[4] == "-" and value[7] == "-":
        value = f"{value}T{'00:00' if day_start else '23:59'}"
    try:
        parsed = datetime.fromisoformat(value)
    except ValueError as exc:
        raise ValueError(f"日期筛选格式不正确：「{value}」") from exc
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=plan_service.get_zone(settings.default_timezone))
    return parsed.astimezone(timezone.utc)


@app.get("/calls")
def calls_page(
    request: Request,
    status: str = "",
    date_from: str = "",
    date_to: str = "",
    customer_id: int | None = None,
    page: int = 1,
    error: str = "",
    notice: str = "",
    db: Session = Depends(get_db),
):
    auth.require_login(request)
    invalid_filter = ""
    try:
        start_bound = _date_filter_bound(date_from, True)
        end_bound = _date_filter_bound(date_to, False)
    except ValueError as exc:
        invalid_filter = str(exc)
        start_bound = end_bound = None

    query = select(CallRecord)
    if status:
        query = query.where(CallRecord.status == status)
    if customer_id:
        query = query.where(CallRecord.customer_id == customer_id)
    if start_bound is not None:
        query = query.where(CallRecord.created_at >= start_bound)
    if end_bound is not None:
        query = query.where(CallRecord.created_at <= end_bound)

    total = db.scalar(select(func.count()).select_from(query.subquery())) or 0
    page = max(1, page)
    last_page = max(1, (total + CALLS_PAGE_SIZE - 1) // CALLS_PAGE_SIZE)
    page = min(page, last_page)
    records = db.scalars(
        query.order_by(CallRecord.created_at.desc())
        .offset((page - 1) * CALLS_PAGE_SIZE)
        .limit(CALLS_PAGE_SIZE)
    ).all()
    # 队列先展示进行中任务（queued/dialing/connected），再按到期时间排终态任务。
    tasks = db.scalars(
        select(CallTask)
        .order_by(
            case((CallTask.status.in_(["queued", "dialing", "connected"]), 0), else_=1),
            CallTask.due_at.asc(),
        )
        .limit(50)
    ).all()
    customers = db.scalars(select(Customer).order_by(Customer.name.asc())).all()
    return render(
        request,
        "calls.html",
        db,
        records=records,
        tasks=tasks,
        customers=customers,
        status_options=CALL_STATUS_OPTIONS,
        status_filter=status,
        date_from=date_from,
        date_to=date_to,
        customer_id=customer_id or 0,
        page=page,
        last_page=last_page,
        total=total,
        page_size=CALLS_PAGE_SIZE,
        error=error or invalid_filter,
        notice=notice,
    )


@app.get("/calls/{record_id}")
def call_detail(
    record_id: int,
    request: Request,
    error: str = "",
    notice: str = "",
    db: Session = Depends(get_db),
):
    auth.require_login(request)
    record = get_or_404(db, CallRecord, record_id)
    events = db.scalars(
        select(CallEvent)
        .where(CallEvent.call_record_id == record.id)
        .order_by(CallEvent.created_at.asc(), CallEvent.id.asc())
    ).all()
    return render(
        request,
        "call_detail.html",
        db,
        record=record,
        events=events,
        error=error,
        notice=notice,
    )


@app.post("/tasks/{task_id}/requeue")
async def task_requeue(task_id: int, request: Request, db: Session = Depends(get_db)):
    """人工重新入队：把终态任务重置为 queued，不新建任务、不改动原计划。"""

    auth.require_login(request)
    task = get_or_404(db, CallTask, task_id)
    form = await request.form()
    next_url = str(form.get("next", "/calls"))
    if not next_url.startswith("/") or next_url.startswith("//"):
        next_url = "/calls"
    try:
        plan_service.requeue_call_task(db, task, message="网页端人工重新入队，等待队列调度。")
    except ValueError as exc:
        db.rollback()
        return redirect_to(f"{next_url}?error={quote(str(exc))}")
    ledger_service.log_action(
        db, "requeue", _user_id(request), "call_task", task.id,
        json.dumps({"action": "manual_requeue"}, ensure_ascii=False),
        _client_ip(request),
    )
    db.commit()
    return redirect_to(f"{next_url}?notice={quote('任务已重新入队，等待 Worker 领取。')}")


@app.post("/calls/{record_id}/feedback")
async def call_feedback(record_id: int, request: Request, db: Session = Depends(get_db)):
    auth.require_login(request)
    record = get_or_404(db, CallRecord, record_id)
    form = await request.form()
    record.operator_feedback = str(form.get("operator_feedback", "")).strip()
    db.commit()
    return redirect_to(f"/calls/{record.id}")
