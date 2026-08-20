from __future__ import annotations

from pathlib import Path

from alembic import command
from alembic.config import Config as AlembicConfig
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session
from app.main import build_ledger_export_workbook
from app.models import BusinessService, NetworkDevice
from app.services.imports import parse_ledger_rows

BASE_DIR = Path(__file__).resolve().parent.parent


def test_flat_export_contains_instructions_and_remains_importable(tmp_path: Path) -> None:
    db_path = tmp_path / "export.db"
    url = f"sqlite:///{db_path}"
    cfg = AlembicConfig(str(BASE_DIR / "alembic.ini"))
    cfg.set_main_option("sqlalchemy.url", url)
    command.upgrade(cfg, "head")
    engine = create_engine(url)
    db = Session(engine)
    service = BusinessService(service_number="848TEST0001", customer_name="测试客户")
    db.add(service)
    db.flush()
    db.add(NetworkDevice(device_code="DEV0001", business_service_id=service.id))
    db.commit()

    workbook = build_ledger_export_workbook(db)
    assert workbook.sheetnames == ["业务设备台账", "填写说明"]
    ledger = workbook["业务设备台账"]
    headers = [cell.value for cell in ledger[2]]
    assert "操作" not in headers
    assert headers[-4:] == ["业务记录ID", "业务版本", "设备记录ID", "设备版本"]
    assert ledger["X1"].value == "系统校验字段（请勿修改）"
    assert ledger["Q2"].comment is not None
    assert len(ledger.data_validations.dataValidation) == 5
    assert workbook["填写说明"].max_row > 10

    exported = tmp_path / "round-trip.xlsx"
    workbook.save(exported)
    rows = parse_ledger_rows(exported)
    assert rows[0]["raw"]["号码"] == "848TEST0001"
    assert rows[0]["raw"]["设备编码"] == "DEV0001"
    db.close()
    engine.dispose()
