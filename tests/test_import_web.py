from __future__ import annotations

import io
import json
from pathlib import Path

import pytest
from alembic import command
from alembic.config import Config as AlembicConfig
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.database import get_db
from app.main import app
from app.models import BusinessService, ChangeSet, ImportBatch, NetworkDevice, StagingRow
from app.services.imports import LEDGER_COLUMNS


ADMIN_PASSWORD = "test-admin-password"
BASE_DIR = Path(__file__).resolve().parent.parent


@pytest.fixture()
def webdb(tmp_path: Path):
    db_path = tmp_path / "web-import.db"
    url = f"sqlite:///{db_path}"
    config = AlembicConfig(str(BASE_DIR / "alembic.ini"))
    config.set_main_option("sqlalchemy.url", url)
    command.upgrade(config, "head")
    engine = create_engine(url, connect_args={"check_same_thread": False})
    factory = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)
    yield factory
    engine.dispose()


def _ledger_row(**overrides: str) -> list[str]:
    values = {
        "号码": "848DIAWEB0001",
        "户名": "Web 测试客户",
        "县分": "汉滨",
        "网格": "汉滨要客",
        "服务状态": "正常开机",
        "入网时间": "20220101",
        "协议到期时间": "20301231",
        "业务类型": "宽带业务",
        "渠道名称": "Web 测试渠道",
        "发展人": "张三",
        "发展人联系电话": "13800000000",
        "客户经理": "李四",
        "客户经理联系电话": "13900000001",
        "网络维护责任人": "王五",
        "网络维护责任人联系电话": "13700000002",
        "设备属性": "资产类",
        "设备编码": "210000WEB01",
        "资产原值或物资购置价格": "1200",
        "设备及物资类型": "光猫",
        "设备厂家+型号": "测试V1",
        "设备放置地点": "机房",
        "设备是否已回收": "否",
        "设备未回收原因": "在用",
    }
    values.update(overrides)
    return [values.get(column, "") for column in LEDGER_COLUMNS]


def _xlsx_bytes(rows: list[list[str]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["业务信息"] + [None] * 12 + ["设备信息"] + [None] * 10)
    sheet.append(LEDGER_COLUMNS)
    for row in rows:
        sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


@pytest.fixture()
def web_client(webdb, monkeypatch, tmp_path: Path):
    def override_get_db():
        db = webdb()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = override_get_db
    # Upload routes intentionally store the original workbook under data/imports.
    # Redirect this test-only storage to a temporary directory.
    import app.main as main_module

    monkeypatch.setattr(main_module, "BASE_DIR", tmp_path)
    (tmp_path / "data" / "imports").mkdir(parents=True)
    client = TestClient(app)
    yield client, webdb
    app.dependency_overrides.pop(get_db, None)


def _login(client: TestClient) -> None:
    response = client.post(
        "/login",
        data={"username": "admin", "password": ADMIN_PASSWORD},
        follow_redirects=False,
    )
    assert response.status_code == 303, response.text


def _upload(client: TestClient, content: bytes):
    response = client.post(
        "/imports/upload",
        files={
            "file": (
                "web-ledger.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        follow_redirects=False,
    )
    assert response.status_code == 303, response.text
    batch_id = int(response.headers["location"].rsplit("/", 1)[1])
    return batch_id


def _row_form(row: list[str]) -> dict[str, str]:
    return dict(zip(LEDGER_COLUMNS, row, strict=True))


def test_upload_and_correct_staging_row_revalidates(web_client) -> None:
    client, webdb = web_client
    _login(client)
    original = _ledger_row(**{"协议到期时间": ""})
    batch_id = _upload(client, _xlsx_bytes([original]))

    with webdb() as db:
        batch = db.get(ImportBatch, batch_id)
        row = db.scalars(select(StagingRow).where(StagingRow.batch_id == batch_id)).one()
        assert batch.status == "ready"
        assert row.status == "missing"
        row_id = row.id
        source_before = json.loads(row.raw_json)

    corrected = _row_form(original)
    corrected["协议到期时间"] = "20301231"
    response = client.post(
        f"/imports/{batch_id}/rows/{row_id}/correct",
        data=corrected,
        follow_redirects=False,
    )
    assert response.status_code == 303, response.text

    with webdb() as db:
        batch = db.get(ImportBatch, batch_id)
        row = db.get(StagingRow, row_id)
        assert batch.status == "ready"
        assert batch.missing_count == 0
        assert row.status == "valid"
        assert json.loads(row.raw_json)["协议到期时间"] == source_before["协议到期时间"] == ""
        assert json.loads(row.mapped_json)["manual_raw"]["协议到期时间"] == "20301231"


def test_business_and_network_reviews_require_self_confirmation_then_apply(web_client) -> None:
    client, webdb = web_client
    _login(client)
    batch_id = _upload(client, _xlsx_bytes([_ledger_row()]))

    submit = client.post(f"/imports/{batch_id}/submit", follow_redirects=False)
    assert submit.status_code == 303, submit.text

    duplicate_submit = client.post(f"/imports/{batch_id}/submit", follow_redirects=False)
    assert duplicate_submit.status_code == 409
    assert "只有校验完成的批次" in duplicate_submit.text
    with webdb() as db:
        change_sets = db.scalars(
            select(ChangeSet).where(ChangeSet.import_batch_id == batch_id).order_by(ChangeSet.domain)
        ).all()
        assert {change.domain for change in change_sets} == {"business", "network"}
        assert all(change.status == "submitted" for change in change_sets)
        business_id = next(change.id for change in change_sets if change.domain == "business")
        network_id = next(change.id for change in change_sets if change.domain == "network")

    # The built-in admin principal is also the importer and must explicitly confirm self-review.
    response = client.post(
        f"/reviews/{business_id}/decision", data={"decision": "approved"}, follow_redirects=False
    )
    assert response.status_code == 400
    assert "二次确认" in response.text

    for change_id in (business_id, network_id):
        response = client.post(
            f"/reviews/{change_id}/decision",
            data={"decision": "approved", "self_review_confirmed": "1"},
            follow_redirects=False,
        )
        assert response.status_code == 303, response.text

    # Applying the network set before business is rejected by the service dependency.
    response = client.post(f"/reviews/{network_id}/apply", follow_redirects=False)
    assert response.status_code == 409
    assert "业务审核" in response.text

    for change_id in (business_id, network_id):
        response = client.post(f"/reviews/{change_id}/apply", follow_redirects=False)
        assert response.status_code == 303, response.text

    with webdb() as db:
        batch = db.get(ImportBatch, batch_id)
        service = db.scalars(
            select(BusinessService).where(BusinessService.service_number == "848DIAWEB0001")
        ).one()
        device = db.scalars(
            select(NetworkDevice).where(NetworkDevice.device_code == "210000WEB01")
        ).one()
        assert batch.status == "applied"
        assert service.customer.name == "Web 测试客户"
        assert device.business_service_id == service.id
        assert {db.get(ChangeSet, business_id).status, db.get(ChangeSet, network_id).status} == {"applied"}

    export = client.get("/exports/ledger.xlsx")
    assert export.status_code == 200
    workbook = load_workbook(io.BytesIO(export.content), read_only=True, data_only=True)
    exported_rows = list(
        workbook["业务设备台账"].iter_rows(min_row=3, values_only=True)
    )
    assert any(row[0] == "848DIAWEB0001" and row[16] == "210000WEB01" for row in exported_rows)
    workbook.close()


def test_flat_ledger_export_is_downloadable(web_client) -> None:
    client, _ = web_client
    _login(client)
    response = client.get("/exports/ledger.xlsx")
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(io.BytesIO(response.content), read_only=True, data_only=True)
    assert workbook.sheetnames[0] == "业务设备台账"
    assert list(workbook["业务设备台账"].iter_rows(min_row=2, max_row=2, values_only=True))[0][:2] == (
        "号码",
        "户名",
    )
    workbook.close()
