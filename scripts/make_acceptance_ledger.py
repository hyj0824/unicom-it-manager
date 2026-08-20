"""Generate the import/export acceptance ledger workbook.

The workbook is deliberately synthetic: it is safe to use in an acceptance
environment and exercises both the valid import path and the validation report.
Running this file again replaces the ignored ``docs/验收测试台账.xlsx`` file.
"""

from __future__ import annotations

import argparse
import sys
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Running a file below ``scripts/`` puts that directory on sys.path, not the
# repository root where the ``app`` package lives.
REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from app.services.imports import (
    LEDGER_COLUMNS,
    TECHNICAL_COLUMNS,
    header_mapping,
    parse_ledger_rows,
)


GROUPS = (
    ("业务信息", 1, 13),
    ("设备信息", 14, 23),
    ("系统校验字段", 24, 27),
)


def _dates(anchor: date) -> dict[str, str]:
    """Return dates in both formats accepted by the import parser."""

    def iso(days: int) -> str:
        return (anchor + timedelta(days=days)).strftime("%Y-%m-%d")

    def compact(days: int) -> str:
        return (anchor + timedelta(days=days)).strftime("%Y%m%d")

    return {
        "past": iso(-30),
        "soon": compact(5),
        "soon_alt": iso(9),
        "medium": iso(60),
        "medium_alt": compact(75),
        "accessed_iso": iso(-400),
        "accessed_compact": compact(-700),
    }


def _service(
    number: str,
    customer: str,
    county: str,
    grid: str,
    status: str,
    accessed: str,
    expires: str,
    business_type: str,
    *,
    channel: str = "政企直销",
    developer: str = "验收发展人",
    developer_phone: str = "13900000001",
    manager: str = "验收客户经理",
    manager_phone: str = "13900000002",
    maintainer: str = "验收网络维护",
    maintainer_phone: str = "13900000003",
) -> dict[str, object]:
    return {
        "号码": number,
        "户名": customer,
        "县分": county,
        "网格": grid,
        "服务状态": status,
        "入网时间": accessed,
        "协议到期时间": expires,
        "业务类型": business_type,
        "渠道名称": channel,
        "发展人": developer,
        "发展人联系电话": developer_phone,
        "客户经理": manager,
        "客户经理联系电话": manager_phone,
        "网络维护责任人": maintainer,
        "网络维护责任人联系电话": maintainer_phone,
    }


def _row(service: dict[str, object], **device: object) -> dict[str, object]:
    values = {column: "" for column in LEDGER_COLUMNS + TECHNICAL_COLUMNS}
    values.update(service)
    # ``设备厂家+型号`` is the public column name, but ``+`` cannot occur in a
    # Python keyword, so row declarations use the local ``厂家型号`` alias.
    if "厂家型号" in device:
        device["设备厂家+型号"] = device.pop("厂家型号")
    values.update(device)
    return values


def build_rows(anchor: date | None = None) -> list[dict[str, object]]:
    """Build 26 rows: 16 valid examples and 10 intentional error examples."""

    dates = _dates(anchor or date.today())
    rows: list[dict[str, object]] = []

    # Each of the three customers owns two services.  Repeated rows for one
    # service keep all 13 business fields identical while changing devices.
    services = {
        "a1": _service(
            "ACC-A-001", "星河智联（验收）", "汉滨", "汉滨要客", "正常开机",
            dates["accessed_iso"], dates["soon"], "数据及网元业务",
        ),
        "a2": _service(
            "ACC-A-002", "星河智联（验收）", "汉滨", "汉滨商企", "正常开机",
            dates["accessed_compact"], dates["medium"], "宽带业务",
        ),
        "b1": _service(
            "ACC-B-001", "远山制造（验收）", "石泉", "石泉政企", "正常开机",
            dates["accessed_iso"], dates["past"], "数据及网元业务",
        ),
        "b2": _service(
            "ACC-B-002", "远山制造（验收）", "石泉", "石泉城区综合网格",
            "主动退网(申请拆机)", dates["accessed_compact"], dates["soon_alt"],
            "宽带业务",
        ),
        "c1": _service(
            "ACC-C-001", "蓝田医院（验收）", "紫阳", "紫阳要客", "正常开机",
            dates["accessed_iso"], dates["soon_alt"], "数据及网元业务",
        ),
        "c2": _service(
            "ACC-C-002", "蓝田医院（验收）", "紫阳", "紫阳城区综合网格",
            "主动退网(申请拆机)", dates["accessed_compact"], dates["medium_alt"],
            "宽带业务",
        ),
    }

    rows.extend(
        [
            _row(services["a1"], 设备属性="资产类", 设备编码="A001、A002", 资产原值或物资购置价格=12000, 设备及物资类型="光猫", 厂家型号="华为 MA5800", 设备放置地点="汉滨机房-A", 设备是否已回收="未回收", 设备未回收原因="在用"),
            _row(services["a1"], 设备属性="成本类", 设备编码="A003", 资产原值或物资购置价格=6800, 设备及物资类型="交换机", 厂家型号="新华三 S5130", 设备放置地点="汉滨机房-B", 设备是否已回收="已回收"),
            _row(services["a1"], 设备属性="资产类", 设备编码="A004", 资产原值或物资购置价格=2300, 设备及物资类型="路由器", 厂家型号="中兴 ZXR10", 设备放置地点="汉滨机房-C", 设备是否已回收="未回收", 设备未回收原因="其他"),
            _row(services["a2"], 设备属性="成本类", 设备编码="A101", 资产原值或物资购置价格=3100, 设备及物资类型="光猫", 厂家型号="烽火 HG", 设备放置地点="汉滨营业厅", 设备是否已回收="已回收"),
            _row(services["a2"], 设备属性="资产类", 设备编码="A102", 资产原值或物资购置价格=4500, 设备及物资类型="交换机", 厂家型号="华为 S1730", 设备放置地点="汉滨营业厅", 设备是否已回收="未回收", 设备未回收原因="遗失"),
            _row(services["b1"], 设备属性="资产类", 设备编码="B001", 资产原值或物资购置价格=9800, 设备及物资类型="路由器", 厂家型号="华为 AR", 设备放置地点="石泉厂区", 设备是否已回收="未回收", 设备未回收原因="纠纷"),
            _row(services["b1"], 设备属性="成本类", 设备编码="B002", 资产原值或物资购置价格=1700, 设备及物资类型="光猫", 厂家型号="中兴 F7607", 设备放置地点="石泉厂区", 设备是否已回收="已回收"),
            _row(services["b2"], 设备属性="资产类", 设备编码="B201,B202", 资产原值或物资购置价格=7600, 设备及物资类型="交换机", 厂家型号="新华三 S5560", 设备放置地点="石泉机房", 设备是否已回收="未回收", 设备未回收原因="在用"),
            _row(services["b2"], 设备属性="成本类", 设备编码="B203", 资产原值或物资购置价格=1900, 设备及物资类型="光猫", 厂家型号="烽火 AN5506", 设备放置地点="石泉机房", 设备是否已回收="未回收", 设备未回收原因="遗失"),
            _row(services["b2"], 设备属性="资产类", 设备编码="B204", 资产原值或物资购置价格=2200, 设备及物资类型="路由器", 厂家型号="中兴 ZXHN", 设备放置地点="石泉机房", 设备是否已回收="已回收"),
            _row(services["c1"], 设备属性="成本类", 设备编码="C001;C002", 资产原值或物资购置价格=5200, 设备及物资类型="光猫", 厂家型号="华为 HG8245", 设备放置地点="紫阳医院机房", 设备是否已回收="未回收", 设备未回收原因="在用"),
            _row(services["c1"], 设备属性="资产类", 设备编码="C003", 资产原值或物资购置价格=8400, 设备及物资类型="交换机", 厂家型号="华为 S5735", 设备放置地点="紫阳医院机房", 设备是否已回收="未回收", 设备未回收原因="其他"),
            _row(services["c1"], 设备属性="成本类", 设备编码="C004", 资产原值或物资购置价格=1350, 设备及物资类型="路由器", 厂家型号="锐捷 RG", 设备放置地点="紫阳医院机房", 设备是否已回收="已回收"),
            _row(services["c2"], 设备属性="资产类", 设备编码="C101", 资产原值或物资购置价格=6300, 设备及物资类型="交换机", 厂家型号="新华三 S1850", 设备放置地点="紫阳营业厅", 设备是否已回收="未回收", 设备未回收原因="纠纷"),
            _row(services["c2"], 设备属性="成本类", 设备编码="C102", 资产原值或物资购置价格=1500, 设备及物资类型="光猫", 厂家型号="中兴 F601", 设备放置地点="紫阳营业厅", 设备是否已回收="已回收"),
            _row(services["c2"], 设备属性="资产类", 设备编码="C103", 资产原值或物资购置价格=4100, 设备及物资类型="路由器", 厂家型号="华为 AR1220", 设备放置地点="紫阳营业厅", 设备是否已回收="未回收", 设备未回收原因="其他"),
        ]
    )

    duplicate = _service(
        "ERR-DUP-001", "重复业务（验收）", "汉滨", "汉滨要客", "正常开机",
        dates["accessed_iso"], dates["medium"], "宽带业务",
    )
    rows.extend(
        [
            _row(duplicate, 设备属性="资产类", 设备编码="DUP001", 资产原值或物资购置价格=1000, 设备及物资类型="光猫", 厂家型号="测试型号", 设备放置地点="测试机房", 设备是否已回收="已回收"),
            _row(duplicate),
            _row(_service("ERR-CONFLICT-001", "冲突业务甲（验收）", "汉滨", "汉滨商企", "正常开机", dates["accessed_iso"], dates["medium"], "宽带业务"), 设备属性="资产类", 设备编码="SHARED-001", 资产原值或物资购置价格=2000, 设备及物资类型="交换机", 厂家型号="测试型号", 设备放置地点="冲突机房", 设备是否已回收="未回收", 设备未回收原因="在用"),
            _row(_service("ERR-CONFLICT-002", "冲突业务乙（验收）", "石泉", "石泉政企", "正常开机", dates["accessed_iso"], dates["medium"], "数据及网元业务"), 设备属性="成本类", 设备编码="SHARED-001", 资产原值或物资购置价格=2000, 设备及物资类型="交换机", 厂家型号="测试型号", 设备放置地点="冲突机房", 设备是否已回收="未回收", 设备未回收原因="纠纷"),
            _row(_service("ERR-ENUM-001", "非法枚举（验收）", "紫阳", "紫阳要客", "非法服务状态", dates["accessed_iso"], dates["medium"], "非法业务类型"), 设备属性="资产类", 设备编码="ENUM/001", 资产原值或物资购置价格=3000, 设备及物资类型="光猫", 厂家型号="测试型号", 设备放置地点="校验机房", 设备是否已回收="已回收"),
            _row(_service("ERR-PHONE-001", "错误电话（验收）", "汉滨", "汉滨要客", "正常开机", dates["accessed_iso"], dates["medium"], "宽带业务", developer_phone="13A00000000"), 设备属性="成本类", 设备编码="PHONE001", 资产原值或物资购置价格=900, 设备及物资类型="光猫", 厂家型号="测试型号", 设备放置地点="校验机房", 设备是否已回收="已回收"),
            _row(_service("", "缺号码（验收）", "汉滨", "汉滨要客", "正常开机", dates["accessed_iso"], dates["medium"], "宽带业务"), 设备属性="资产类", 设备编码="MISSING-NUM", 资产原值或物资购置价格=800, 设备及物资类型="光猫", 厂家型号="测试型号", 设备放置地点="校验机房", 设备是否已回收="已回收"),
            _row(_service("ERR-NAME-001", "", "石泉", "石泉政企", "正常开机", dates["accessed_iso"], dates["medium"], "宽带业务"), 设备属性="资产类", 设备编码="MISSING-NAME", 资产原值或物资购置价格=800, 设备及物资类型="光猫", 厂家型号="测试型号", 设备放置地点="校验机房", 设备是否已回收="已回收"),
            _row(_service("ERR-MISSING-001", "缺协议和经理电话（验收）", "紫阳", "紫阳要客", "正常开机", dates["accessed_iso"], "", "数据及网元业务", manager_phone=""), 设备属性="成本类", 设备编码="MISSING-FIELD", 资产原值或物资购置价格=1100, 设备及物资类型="交换机", 厂家型号="测试型号", 设备放置地点="校验机房", 设备是否已回收="未回收", 设备未回收原因="在用"),
            _row(_service("ERR-SEPARATOR-001", "非法分隔符（验收）", "汉滨", "汉滨商企", "正常开机", dates["accessed_iso"], dates["medium"], "宽带业务"), 设备属性="资产类", 设备编码="BAD/SEP|01", 资产原值或物资购置价格=1200, 设备及物资类型="路由器", 厂家型号="测试型号", 设备放置地点="校验机房", 设备是否已回收="未回收", 设备未回收原因="其他"),
        ]
    )
    return rows


def write_workbook(path: Path, rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "验收台账"

    for title, start, end in GROUPS:
        sheet.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        cell = sheet.cell(1, start, title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for column in range(start, end + 1):
            sheet.cell(1, column).fill = PatternFill("solid", fgColor="1F4E78")

    headers = LEDGER_COLUMNS + TECHNICAL_COLUMNS
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(2, column, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="5B9BD5")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_number, row in enumerate(rows, start=3):
        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(row_number, column, row[header])
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        "号码": 17, "户名": 20, "县分": 10, "网格": 22, "服务状态": 19,
        "入网时间": 13, "协议到期时间": 15, "业务类型": 17, "渠道名称": 14,
        "发展人": 14, "发展人联系电话": 18, "客户经理": 15, "客户经理联系电话": 19,
        "网络维护责任人": 17, "网络维护责任人联系电话": 22, "设备属性": 11,
        "设备编码": 18, "资产原值或物资购置价格": 18, "设备及物资类型": 15,
        "设备厂家+型号": 20, "设备放置地点": 19, "设备是否已回收": 14,
        "设备未回收原因": 14, "业务记录ID": 13, "业务版本": 11, "设备记录ID": 13,
        "设备版本": 11,
    }
    for column, header in enumerate(headers, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = widths.get(header, 14)
    sheet.row_dimensions[1].height = 24
    sheet.row_dimensions[2].height = 38
    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{len(rows) + 2}"
    workbook.save(path)


def verify(path: Path, expected_rows: int) -> None:
    """Re-read the saved file through the production parser and report mapping."""

    try:
        parsed = parse_ledger_rows(path)
        mapping = header_mapping(path)
        workbook = load_workbook(path, read_only=True, data_only=True)
        headers = list(next(workbook.active.iter_rows(min_row=2, max_row=2, values_only=True)))
        workbook.close()
    except Exception as exc:  # pragma: no cover - a clear CLI failure is desired
        print(f"解析异常: 1 ({exc})")
        raise

    expected_headers = LEDGER_COLUMNS + TECHNICAL_COLUMNS
    if headers != expected_headers:
        raise RuntimeError("第 2 行表头顺序与 LEDGER_COLUMNS/TECHNICAL_COLUMNS 不一致")
    if len(parsed) != expected_rows:
        raise RuntimeError(f"解析行数不符：期望 {expected_rows}，实际 {len(parsed)}")
    if any(set(item["raw"]) != set(expected_headers) for item in parsed):
        raise RuntimeError("存在未返回完整列映射的解析行")

    full_mapping = {**mapping, **{column: column for column in TECHNICAL_COLUMNS}}
    print(f"文件: {path}")
    print(f"数据行数: {len(parsed)}")
    print("列映射: " + "、".join(f"{column}←{full_mapping[column]}" for column in expected_headers))
    print("解析异常: 0")


def main() -> None:
    parser = argparse.ArgumentParser(description="生成导入导出字段验收测试台账")
    parser.add_argument(
        "--output",
        type=Path,
        default=REPO_ROOT / "docs" / "验收测试台账.xlsx",
        help="输出 xlsx 路径（默认 docs/验收测试台账.xlsx）",
    )
    args = parser.parse_args()
    rows = build_rows()
    write_workbook(args.output, rows)
    verify(args.output, len(rows))


if __name__ == "__main__":
    main()
